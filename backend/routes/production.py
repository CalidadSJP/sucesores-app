from flask import Blueprint, request, jsonify
from core.database import get_db_connection
from openpyxl import Workbook
from io import BytesIO
from flask import request, jsonify, send_file, current_app
from psycopg2.extras import RealDictCursor
from datetime import date, datetime
from openpyxl.styles import Border, Side, Alignment, Font


production_bp = Blueprint("production", __name__)


# PAGINA | REVISION DE PIEZAS LINEAS DE PRODUCCIÓN

@production_bp.route('/submit-piece-inspection', methods=['POST'])
def submit_piece_inspection():
    try:
        data = request.get_json()

        # Reemplazar campos vacíos con "0" o "" según corresponda
        cleaned_data = {
            'line_id': data.get('line_id') or '0',
            'piece': data.get('piece') or '',
            'inspection_date': data.get('inspection_date') or '0',
            'mesh': data.get('mesh') or '',
            'screw': data.get('screw') or '',
            'bulb_plastic': data.get('bulb_plastic') or '',
            'connector_mounting': data.get('connector_mounting') or '',
            'connector_adjustment': data.get('connector_adjustment') or '',
            'connector_plastic': data.get('connector_plastic') or '',
            'transducer_adjustment': data.get('transducer_adjustment') or '',
            'probe_plastic': data.get('probe_plastic') or '',
            'observations': data.get('observations') if data.get('observations') not in [None, ''] else ''
        }

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO production_parts_inspection (
                line_id, piece, inspection_date, mesh, screw,
                bulb_plastic, connector_mounting, connector_adjustment,
                connector_plastic, transducer_adjustment, probe_plastic,
                observations
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            cleaned_data['line_id'],
            cleaned_data['piece'],
            cleaned_data['inspection_date'],
            cleaned_data['mesh'],
            cleaned_data['screw'],
            cleaned_data['bulb_plastic'],
            cleaned_data['connector_mounting'],
            cleaned_data['connector_adjustment'],
            cleaned_data['connector_plastic'],
            cleaned_data['transducer_adjustment'],
            cleaned_data['probe_plastic'],
            cleaned_data['observations']
        ))

        conn.commit()
        return jsonify({'message': 'Inspección registrada correctamente'}), 201

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'error': 'Error al guardar la inspección', 'details': str(e)}), 500

    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

@production_bp.route('/get-piece-inspections', methods=['GET'])
def get_piece_inspections():
    line_id = request.args.get('line_id')
    piece = request.args.get('piece')
    date = request.args.get('date')

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        query = """
            SELECT i.id, i.inspection_date, i.piece, pl.line_name,
                   i.mesh, i.screw, i.bulb_plastic, i.connector_mounting,
                   i.connector_adjustment, i.connector_plastic,
                   i.transducer_adjustment, i.probe_plastic, i.observations
            FROM production_parts_inspection i
            JOIN production_line pl ON i.line_id = pl.id
            WHERE 1=1
        """
        params = []

        if line_id:
            query += " AND i.line_id = %s"
            params.append(line_id)
        if piece:
            query += " AND i.piece ILIKE %s"
            params.append(f"%{piece}%")
        if date:
            query += " AND i.inspection_date = %s"
            params.append(date)

        query += " ORDER BY i.inspection_date DESC"

        cur.execute(query, params)
        records = cur.fetchall()

        cur.close()
        conn.close()

        return jsonify(records), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_bp.route('/download-piece-inspections', methods=['GET'])
def download_piece_inspections_excel():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        query = '''
            SELECT
                pi.inspection_date,
                pl.line_name,
                pi.piece,
                pi.mesh,
                pi.screw,
                pi.bulb_plastic,
                pi.connector_mounting,
                pi.connector_adjustment,
                pi.connector_plastic,
                pi.transducer_adjustment,
                pi.probe_plastic,
                pi.observations
            FROM production_parts_inspection pi
            JOIN production_line pl ON pi.line_id = pl.id
            ORDER BY pi.inspection_date DESC
        '''
        cur.execute(query)
        data = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Inspecciones Piezas"

        # Definir estilos de borde y centrado
        thick_border = Border(right=Side(style='thick'))
        thin_border = Border(right=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold_font = Font(bold=True)

        # Crear cabecera de secciones (Row 1)
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:B2')
        ws.merge_cells('C1:C2')
        ws.merge_cells('D1:F1')
        ws.merge_cells('G1:I1')
        ws.merge_cells('J1:K1')
        ws.merge_cells('L1:L2')

        ws['A1'] = 'Fecha'
        ws['B1'] = 'Línea'
        ws['C1'] = 'Pieza'
        ws['D1'] = 'Revisión de Bulbo'
        ws['G1'] = 'Revisión del Conector de Alimentación'
        ws['J1'] = 'Revisión del Conector Sonda'
        ws['L1'] = 'Observaciones'

        # Sub-encabezados (Row 2)
        headers_row2 = [
            'Malla', 'Tornillo', 'Plástico',
            'Sujeción', 'Ajuste (Tuerca)', 'Plástico',
            'Ajuste al Transductor', 'Plástico'
        ]
        start_col = 4  # Columna D
        for idx, header in enumerate(headers_row2):
            cell = ws.cell(row=2, column=start_col + idx, value=header)
            cell.alignment = center_alignment
            cell.font = bold_font

        # Aplicar estilos a la fila 1 (Secciones)
        for col in range(1, 13 + 1):
            cell = ws.cell(row=1, column=col)
            cell.alignment = center_alignment
            cell.font = bold_font
            # Bordes gruesos para separación visual
            if col in [3, 6, 9, 11, 12]:  # Líneas de separación visual
                cell.border = thick_border
            else:
                cell.border = thin_border

        # Aplicar bordes a la fila 2
        for col in range(4, 12):
            cell = ws.cell(row=2, column=col)
            if col in [6, 9, 11]:  # Líneas de separación visual
                cell.border = thick_border
            else:
                cell.border = thin_border

        # Agregar los datos
        def safe_str(value):
            if isinstance(value, (datetime, date)):
                return value.strftime('%d-%m-%Y')
            return value

        data_start_row = 3
        for row_idx, row_data in enumerate(data, start=data_start_row):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=safe_str(cell_value))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Bordes de separación
                if col_idx in [3, 6, 9, 11, 12]:  # Columnas de separación visual
                    cell.border = thick_border
                else:
                    cell.border = thin_border

        # Ajustar anchos de columnas (Opcional)
        column_widths = {
            'A': 12, 'B': 15, 'C': 15, 'D': 10, 'E': 10, 'F': 12,
            'G': 12, 'H': 14, 'I': 12, 'J': 18, 'K': 12, 'L': 25
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        cur.close()
        conn.close()

        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="inspecciones_produccion.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

