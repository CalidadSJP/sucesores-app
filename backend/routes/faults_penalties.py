from flask import Blueprint, request, jsonify
from core.database import get_db_connection
from openpyxl import Workbook
from io import BytesIO
from flask import request, jsonify, send_file, current_app
from psycopg2.extras import RealDictCursor
from datetime import date, datetime
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


penailties_bp = Blueprint("faults_penalties", __name__)

#PAGINA | FALTAS Y MULTAS

@penailties_bp.route('/faults', methods=['POST'])
def add_fault():
    try:
        data = request.get_json()

        # Validaciones mínimas obligatorias
        required_fields = ['personnel_id', 'description', 'responsible', 'date', 'fault_type_id', 'severity']
        for field in required_fields:
            if field not in data or not data[field] and data[field] != 0:
                return jsonify({'error': f'Campo obligatorio faltante: {field}'}), 400

        # --- Casteos seguros ---
        try:
            personnel_id = int(data['personnel_id'])
        except Exception:
            return jsonify({'error': f"personnel_id inválido: {data.get('personnel_id')}"}), 400

        description = str(data.get('description', '')).strip().upper()
        responsible = str(data.get('responsible', '')).strip().upper()

        fault_type_id = data.get('fault_type_id')
        try:
            fault_type_id = int(fault_type_id) if fault_type_id not in (None, '', []) else None
        except Exception:
            return jsonify({'error': f'fault_type_id inválido: {fault_type_id}'}), 400

        severity = str(data.get('severity', '')).strip().upper()

        penalty_description = str(data.get('penalty_description', '')).strip().upper() if data.get('penalty_description') else ''
        numeration = data.get('numeration')
        try:
            numeration = int(numeration) if numeration not in (None, '', []) else None
        except Exception:
            return jsonify({'error': f'numeration inválido: {numeration}'}), 400

        date = str(data.get('date'))

        # Debug de los valores de entrada
        print("[DEBUG] Datos recibidos:", data)
        print("[DEBUG] Parsed -> personnel_id:", personnel_id,
              "| description:", description,
              "| responsible:", responsible,
              "| date:", date,
              "| fault_type_id:", fault_type_id,
              "| severity:", severity,
              "| penalty_description:", penalty_description,
              "| numeration:", numeration)

        # Validar severidad permitida
        if severity not in ['LEVE', 'GRAVE', 'MUY GRAVE']:
            return jsonify({'error': f'Severidad inválida: {severity}'}), 400

        conn = get_db_connection()
        cur = conn.cursor()

        # Insertar la falta
        cur.execute("""
            INSERT INTO faults (personnel_id, description, responsible, date, fault_type_id, severity)
            VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
        """, (personnel_id, description, responsible, date, fault_type_id, severity))
        fault_id = cur.fetchone()[0]
        print(f"[DEBUG] Falta insertada con ID: {fault_id}")

        penalty_id = None
        # --- Lógica de multas ---
        # 1) Si es GRAVE o MUY GRAVE Y viene penalty_description -> crear multa
        # 2) Si es LEVE y viene penalty_description -> crear multa (no contará para acumulación porque la asociamos)
        if (severity in ['GRAVE', 'MUY GRAVE'] or severity == 'LEVE') and penalty_description:
            # Insertar la multa
            cur.execute("""
                INSERT INTO penalties (personnel_id, description, responsible, date, fault_type_id, numeration, fault_description)
                VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
            """, (
                personnel_id,
                penalty_description,
                responsible,
                date,
                fault_type_id,
                numeration,
                description  # guardamos la descripción de la falta original en fault_description
            ))
            penalty_id = cur.fetchone()[0]
            print(f"[DEBUG] Multa insertada con ID: {penalty_id}")

            # Asociar la falta actual con la multa (para que no cuente como "no penalizada")
            cur.execute("""
                INSERT INTO penalties_faults (penalty_id, fault_id)
                VALUES (%s, %s)
            """, (penalty_id, fault_id))
            print(f"[DEBUG] Asociada falta {fault_id} con multa {penalty_id}")

        # --- Multa automática por acumulación de 5 faltas LEVES (solo contar faltas LEVES sin penalizar) ---
        cur.execute("""
            SELECT f.id, f.description FROM faults f
            LEFT JOIN penalties_faults pf ON f.id = pf.fault_id
            WHERE f.personnel_id = %s AND pf.fault_id IS NULL AND f.severity = 'LEVE'
            ORDER BY f.date ASC, f.id ASC
            LIMIT 5
        """, (personnel_id,))
        leves = cur.fetchall()
        print(f"[DEBUG] Faltas leves no penalizadas encontradas: {len(leves)}")
        print(f"[DEBUG] IDs y descripciones: {leves}")

        if len(leves) == 5:
            try:
                # Insertar multa automática por acumulación
                cur.execute("""
                    INSERT INTO penalties (personnel_id, description, responsible, date, fault_type_id, numeration, fault_description)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    personnel_id,
                    'MULTA POR ACUMULACIÓN DE 5 FALTAS LEVES',
                    responsible,
                    date,
                    7,   # si quieres usar un tipo específico para esta multa
                    None,
                    ' / '.join([l[1] for l in leves])
                ))
                auto_penalty_id = cur.fetchone()[0]
                print(f"[DEBUG] Multa automática por 5 faltas leves creada con ID: {auto_penalty_id}")

                # Asociar faltas a la multa automática
                for l in leves:
                    cur.execute("""
                        INSERT INTO penalties_faults (penalty_id, fault_id)
                        VALUES (%s, %s)
                    """, (auto_penalty_id, l[0]))
                print(f"[DEBUG] Faltas asociadas con multa ID {auto_penalty_id}: {[l[0] for l in leves]}")

            except Exception as e:
                print(f"[ERROR] Falló la creación o asociación de multa automática: {e}")
                conn.rollback()
                raise

        # Contar faltas LEVES no penalizadas restantes (sólo estas afectan el ahorcado)
        cur.execute("""
            SELECT COUNT(*) FROM faults f
            LEFT JOIN penalties_faults pf ON f.id = pf.fault_id
            WHERE f.personnel_id = %s AND pf.fault_id IS NULL AND f.severity = 'LEVE'
        """, (personnel_id,))
        active_faults = cur.fetchone()[0]

        print(f"[DEBUG] Faltas LEVES activas no penalizadas restantes: {active_faults}")

        conn.commit()
        cur.close()
        conn.close()

        # Respuesta: incluimos penalty_id (si se creó) para que el frontend pueda mostrar la multa inmediatamente
        response = {'status': 'success', 'fault_id': fault_id, 'active_faults': active_faults}
        if penalty_id:
            response['penalty_id'] = penalty_id

        return jsonify(response)

    except Exception as e:
        import traceback
        print("[ERROR] Excepción en /faults:", traceback.format_exc())
        return jsonify({'error': 'Error interno en el servidor', 'details': str(e)}), 500

@penailties_bp.route('/faults/<int:personnel_id>', methods=['GET']) 
def get_faults(personnel_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT f.id,
               f.date,
               f.description,
               'falta' AS type_color,
               f.responsible,
               ft.name AS type,
               UPPER(TRIM(f.severity)) AS severity,
               CASE WHEN pf.fault_id IS NULL THEN FALSE ELSE TRUE END AS penalized,
               pen.id,
               pen.date,
               pen.description,
               pen.responsible,
               pen.fault_type_id,
               pen.numeration,
               pen.fault_description
        FROM faults f
        LEFT JOIN fault_types ft ON f.fault_type_id = ft.id
        LEFT JOIN penalties_faults pf ON f.id = pf.fault_id
        LEFT JOIN penalties pen ON pf.penalty_id = pen.id
        WHERE f.personnel_id = %s
        ORDER BY f.date DESC, f.id DESC
    """, (personnel_id,))
    rows = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*)
        FROM faults f
        LEFT JOIN penalties_faults pf ON f.id = pf.fault_id
        WHERE f.personnel_id = %s
          AND pf.fault_id IS NULL
          AND UPPER(TRIM(f.severity)) = 'LEVE'
    """, (personnel_id,))
    active_faults = cur.fetchone()[0]

    cur.close()
    conn.close()

    faults = []
    for r in rows:
        fault = {
            'id'         : r[0],
            'date'       : r[1].isoformat(),
            'description': r[2],
            'type_color' : r[3],
            'responsible': r[4],
            'type'       : r[5],
            'severity'   : r[6],
            'penalized'  : r[7]
        }
        if r[8]:  # Si tiene multa asociada
            fault['penalty'] = {
                'id': r[8],
                'date': r[9].isoformat() if r[9] else None,
                'description': r[10],
                'responsible': r[11],
                'fault_type_id': r[12],
                'numeration': r[13],
                'fault_description': r[14]
            }
        faults.append(fault)

    return jsonify({'faults': faults, 'active_faults': active_faults})

@penailties_bp.route('/get-personnel-list', methods=['GET'])
def get_personnel_list():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT id, name FROM personnel ORDER by name ASC')  # Solo campos relevantes

        rows = cur.fetchall()
        personnel = [{"id": row[0], "name": row[1]} for row in rows]

        cur.close()
        conn.close()

        return jsonify({"personnel": personnel}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@penailties_bp.route('/download-faults-penalties', methods=['GET'])
def download_faults_and_penalties():
    conn = get_db_connection()
    cursor = conn.cursor()

    # === FALTAS ===
    cursor.execute("""
        SELECT 
            f.date,
            f.description,
            f.responsible,
            f.severity,
            ft.name AS fault_type_name,
            p.name AS personnel_name,
            p.identifier
        FROM faults f
        JOIN personnel p ON f.personnel_id = p.id
        LEFT JOIN fault_types ft ON f.fault_type_id = ft.id
        ORDER BY f.date DESC
    """)
    faults = cursor.fetchall()

    # === MULTAS ===
    cursor.execute("""
        SELECT 
            pen.date,
            pen.fault_description,
            pen.description,
            pen.responsible,
            pen.numeration,
            ft.name AS fault_type_name,
            p.name AS personnel_name,
            p.identifier
        FROM penalties pen
        JOIN personnel p ON pen.personnel_id = p.id
        LEFT JOIN fault_types ft ON pen.fault_type_id = ft.id
        ORDER BY pen.date DESC
    """)
    penalties = cursor.fetchall()

    cursor.close()
    conn.close()

    # === Crear archivo Excel ===
    workbook = Workbook()

    # --- Estilos ---
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def format_sheet(sheet, headers, data, description_cols=None):
        # Cabeceras
        sheet.append(headers)
        for col_num, col in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

        # Filas
        for row in data:
            sheet.append(row)
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=sheet.max_row, column=col_num)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = thin_border

        # Ajustar ancho de columnas
        for col_num, col in enumerate(headers, 1):
            max_length = max(len(str(sheet.cell(row=row, column=col_num).value or "")) for row in range(1, sheet.max_row + 1))
            if description_cols and col_num in description_cols:
                adjusted_width = min(40, max_length + 2)  # limitar descripciones
            else:
                adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # --- Hoja de Faltas ---
    faults_sheet = workbook.active
    faults_sheet.title = "Faltas"
    faults_data = [
        [
            row[6],  # cedula
            row[5],  # empleado
            row[0].strftime("%d/%m/%Y") if row[0] else "",
            row[1],  # descripción
            row[2],  # responsable
            row[4],  # tipo
            row[3]   # severidad
        ] for row in faults
    ]
    format_sheet(
        faults_sheet,
        ["Cédula", "Empleado", "Fecha", "Descripción", "Responsable", "Tipo", "Severidad"],
        faults_data,
        description_cols=[4]  # columna "Descripción"
    )

    # --- Hoja de Multas ---
    penalties_sheet = workbook.create_sheet("Multas")
    penalties_data = [
        [
            row[7],  # cedula
            row[6],  # empleado
            row[0].strftime("%d/%m/%Y") if row[0] else "",
            row[1],  # descripción falta
            row[2],  # descripción multa
            row[3],  # responsable
            row[4],  # numeración
            row[5]   # tipo
        ] for row in penalties
    ]
    format_sheet(
        penalties_sheet,
        ["Cédula", "Empleado", "Fecha", "Descripción Falta", "Descripción Multa", "Responsable", "Numeración", "Tipo"],
        penalties_data,
        description_cols=[4, 5]  # columnas de descripciones
    )

    # === Guardar en memoria y enviar ===
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="faltas_y_multas.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@penailties_bp.route('/get-fault-types')
def get_fault_types():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM fault_types ORDER BY id")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify({'types': [{'id': row[0], 'name': row[1]} for row in rows]})


#PAGINA | LISTADO DE FALTAS Y MULTAS

@penailties_bp.route('/get-faults', methods=['GET'])
def get_faults_list():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            f.id,
            f.date,
            f.description,
            f.responsible,
            f.severity,
            ft.name AS fault_type_name,
            p.name AS personnel_name
        FROM faults f
        JOIN personnel p ON f.personnel_id = p.id
        LEFT JOIN fault_types ft ON f.fault_type_id = ft.id
        ORDER BY f.date DESC
    """)
    rows = cursor.fetchall()
    conn.close()
    faults = [
        {
            'id': row[0],
            'date': row[1].isoformat(),
            'full_name': row[6],
            'description': row[2],
            'responsible': row[3],
            'type': row[5],  # nombre del tipo de falta o multa
            'severity': row[4],
        } for row in rows
    ]
    return jsonify(faults)

@penailties_bp.route('/get-penalties', methods=['GET'])
def get_penalties():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            pen.id,
            pen.date,
            pen.fault_description,
            pen.description,
            pen.responsible,
            pen.numeration,
            ft.name AS fault_type_name,
            p.name AS personnel_name
        FROM penalties pen
        JOIN personnel p ON pen.personnel_id = p.id
        LEFT JOIN fault_types ft ON pen.fault_type_id = ft.id
        ORDER BY pen.date DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    penalties = [
        {
            'id': row[0],
            'date': row[1].isoformat(),
            'fault_description': row[2],
            'description': row[3],
            'responsible': row[4],
            'numeration': row[5],
            'type': row[6],  # nombre del tipo de falta o multa
            'full_name': row[7],
        } for row in rows
    ]
    return jsonify(penalties)

# PAGINA | REVISION DE PIEZAS LINEAS DE PRODUCCIÓN

@penailties_bp.route('/submit-piece-inspection', methods=['POST'])
def submit_piece_inspection():
    try:
        data = request.get_json()

        # Reemplazar campos vacíos con "0" o "" según corresponda
        cleaned_data = {
            'line_id': data.get('line_id') or '0',
            'piece': data.get('piece') or '',
            'inspection_date': data.get('inspection_date') or '0',
            'mesh': data.get('mesh') or '',
            'screw': data.get('screw') or '',
            'bulb_plastic': data.get('bulb_plastic') or '',
            'connector_mounting': data.get('connector_mounting') or '',
            'connector_adjustment': data.get('connector_adjustment') or '',
            'connector_plastic': data.get('connector_plastic') or '',
            'transducer_adjustment': data.get('transducer_adjustment') or '',
            'probe_plastic': data.get('probe_plastic') or '',
            'observations': data.get('observations') if data.get('observations') not in [None, ''] else ''
        }

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO production_parts_inspection (
                line_id, piece, inspection_date, mesh, screw,
                bulb_plastic, connector_mounting, connector_adjustment,
                connector_plastic, transducer_adjustment, probe_plastic,
                observations
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            cleaned_data['line_id'],
            cleaned_data['piece'],
            cleaned_data['inspection_date'],
            cleaned_data['mesh'],
            cleaned_data['screw'],
            cleaned_data['bulb_plastic'],
            cleaned_data['connector_mounting'],
            cleaned_data['connector_adjustment'],
            cleaned_data['connector_plastic'],
            cleaned_data['transducer_adjustment'],
            cleaned_data['probe_plastic'],
            cleaned_data['observations']
        ))

        conn.commit()
        return jsonify({'message': 'Inspección registrada correctamente'}), 201

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'error': 'Error al guardar la inspección', 'details': str(e)}), 500

    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

@penailties_bp.route('/get-piece-inspections', methods=['GET'])
def get_piece_inspections():
    line_id = request.args.get('line_id')
    piece = request.args.get('piece')
    date = request.args.get('date')

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        query = """
            SELECT i.id, i.inspection_date, i.piece, pl.line_name,
                   i.mesh, i.screw, i.bulb_plastic, i.connector_mounting,
                   i.connector_adjustment, i.connector_plastic,
                   i.transducer_adjustment, i.probe_plastic, i.observations
            FROM production_parts_inspection i
            JOIN production_line pl ON i.line_id = pl.id
            WHERE 1=1
        """
        params = []

        if line_id:
            query += " AND i.line_id = %s"
            params.append(line_id)
        if piece:
            query += " AND i.piece ILIKE %s"
            params.append(f"%{piece}%")
        if date:
            query += " AND i.inspection_date = %s"
            params.append(date)

        query += " ORDER BY i.inspection_date DESC"

        cur.execute(query, params)
        records = cur.fetchall()

        cur.close()
        conn.close()

        return jsonify(records), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@penailties_bp.route('/download-piece-inspections', methods=['GET'])
def download_piece_inspections_excel():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        query = '''
            SELECT
                pi.inspection_date,
                pl.line_name,
                pi.piece,
                pi.mesh,
                pi.screw,
                pi.bulb_plastic,
                pi.connector_mounting,
                pi.connector_adjustment,
                pi.connector_plastic,
                pi.transducer_adjustment,
                pi.probe_plastic,
                pi.observations
            FROM production_parts_inspection pi
            JOIN production_line pl ON pi.line_id = pl.id
            ORDER BY pi.inspection_date DESC
        '''
        cur.execute(query)
        data = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Inspecciones Piezas"

        # Definir estilos de borde y centrado
        thick_border = Border(right=Side(style='thick'))
        thin_border = Border(right=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold_font = Font(bold=True)

        # Crear cabecera de secciones (Row 1)
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:B2')
        ws.merge_cells('C1:C2')
        ws.merge_cells('D1:F1')
        ws.merge_cells('G1:I1')
        ws.merge_cells('J1:K1')
        ws.merge_cells('L1:L2')

        ws['A1'] = 'Fecha'
        ws['B1'] = 'Línea'
        ws['C1'] = 'Pieza'
        ws['D1'] = 'Revisión de Bulbo'
        ws['G1'] = 'Revisión del Conector de Alimentación'
        ws['J1'] = 'Revisión del Conector Sonda'
        ws['L1'] = 'Observaciones'

        # Sub-encabezados (Row 2)
        headers_row2 = [
            'Malla', 'Tornillo', 'Plástico',
            'Sujeción', 'Ajuste (Tuerca)', 'Plástico',
            'Ajuste al Transductor', 'Plástico'
        ]
        start_col = 4  # Columna D
        for idx, header in enumerate(headers_row2):
            cell = ws.cell(row=2, column=start_col + idx, value=header)
            cell.alignment = center_alignment
            cell.font = bold_font

        # Aplicar estilos a la fila 1 (Secciones)
        for col in range(1, 13 + 1):
            cell = ws.cell(row=1, column=col)
            cell.alignment = center_alignment
            cell.font = bold_font
            # Bordes gruesos para separación visual
            if col in [3, 6, 9, 11, 12]:  # Líneas de separación visual
                cell.border = thick_border
            else:
                cell.border = thin_border

        # Aplicar bordes a la fila 2
        for col in range(4, 12):
            cell = ws.cell(row=2, column=col)
            if col in [6, 9, 11]:  # Líneas de separación visual
                cell.border = thick_border
            else:
                cell.border = thin_border

        # Agregar los datos
        def safe_str(value):
            if isinstance(value, (datetime, date)):
                return value.strftime('%d-%m-%Y')
            return value

        data_start_row = 3
        for row_idx, row_data in enumerate(data, start=data_start_row):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=safe_str(cell_value))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Bordes de separación
                if col_idx in [3, 6, 9, 11, 12]:  # Columnas de separación visual
                    cell.border = thick_border
                else:
                    cell.border = thin_border

        # Ajustar anchos de columnas (Opcional)
        column_widths = {
            'A': 12, 'B': 15, 'C': 15, 'D': 10, 'E': 10, 'F': 12,
            'G': 12, 'H': 14, 'I': 12, 'J': 18, 'K': 12, 'L': 25
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        cur.close()
        conn.close()

        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="inspecciones_produccion.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

