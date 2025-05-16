from dotenv import load_dotenv
import os
import psycopg2
from psycopg2.extras import RealDictCursor
from flask import Flask, request, jsonify, render_template, send_file, send_from_directory, abort
from flask_cors import CORS
from io import BytesIO
from openpyxl import Workbook
from werkzeug.utils import secure_filename
from datetime import datetime, date, time
from flask import send_from_directory
import numpy as np

#Reiniciar servicio: httpd.exe -k restart -n "sucesores-app"

# Cargar las variables de entorno
load_dotenv()

# Crear la instancia de Flask
app = Flask(__name__, static_folder='static', template_folder='templates')

CORS(app, resources={r"/*": {"origins": ["http://192.168.0.251:8080"]}})

# Conexión a PostgreSQL
def get_db_connection():
    conn = psycopg2.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        database=os.getenv('DB_NAME'),
        user=os.getenv('DB_USER'),
        password=os.getenv('DB_PASSWORD'),
    )
    return conn

# Configurar la carpeta de carga
UPLOAD_FOLDER = 'D:/Projects/sucesores-app-data/Ingreso a Bodega de Aditivos'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Limitar el tamaño máximo del archivo a 16 MB

UPLOAD_MATERIAL_FOLDER = 'D:/Projects/sucesores-app-data/Ingreso de Material de Empaque'
app.config['UPLOAD_MATERIAL_FOLDER'] = UPLOAD_MATERIAL_FOLDER

@app.route('/')
def index():
    return render_template('index.html')

#PAGINA FORMULARIO DEL PERSONAL

@app.route('/submit-form', methods=['POST']) #Método para subir el formulario
def submit_form():
    try:
        data = request.json
        print(f"Datos recibidos: {data}")  # Log para ver los datos recibidos

        # Establecer conexión con la base de datos
        conn = get_db_connection()
        cur = conn.cursor()

        # Reemplazar valores nulos en los datos si es necesario (por ejemplo, si 'observaciones' puede ser nulo)
        observaciones = data.get('observaciones', None)

        # Ejecutar consulta SQL para insertar datos
        cur.execute(''' 
            INSERT INTO inspection 
            (fecha, turno, area, nombre_operario, manos_limpias, uniforme_limpio, no_objetos_personales, 
             heridas_protegidas, cofia_bien_puesta, mascarilla_bien_colocada, protector_auditivo, 
             unas_cortas, guantes_limpios, pestanas, barba_bigote, medicamento_autorizado, supervisor, observaciones)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            data['fecha'], data['turno'], data['area'], data['nombre_operario'], data['manos_limpias'], 
            data['uniforme_limpio'], data['no_objetos_personales'], data['heridas_protegidas'], 
            data['cofia_bien_puesta'], data['mascarilla_bien_colocada'], data['protector_auditivo'], 
            data['unas_cortas'], data['guantes_limpios'], data['pestanas'], data['barba_bigote'], 
            data['medicamento_autorizado'], data['supervisor'], observaciones  # Observaciones se maneja para ser None si está vacío
        ))

        # Confirmar los cambios en la base de datos
        conn.commit()

        # Cerrar cursor y conexión
        cur.close()
        conn.close()

        return jsonify({"message": "Formulario guardado en la base de datos."}), 200

    except Exception as e:
        print(f"Error general: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/download-inspection', methods=['GET']) #Método para descargar el registro de inpección del personal
def download_inspection():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Seleccionar solo las columnas que deseas incluir en el Excel
        cur.execute('''
            SELECT fecha, turno, area, nombre_operario, manos_limpias, uniforme_limpio, 
                   no_objetos_personales, heridas_protegidas, cofia_bien_puesta, 
                   mascarilla_bien_colocada, protector_auditivo, unas_cortas, 
                   guantes_limpios, pestanas, barba_bigote, medicamento_autorizado, 
                   supervisor, observaciones
            FROM inspection
            ORDER BY fecha DESC
        ''')
        data = cur.fetchall()

        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspecciones"

        # Agregar encabezados
        headers = ["Fecha", "Turno", "Área", "Nombre del operario", "Manos limpias", "Uniforme limpio", 
                   "No objetos personales", "Heridas protegidas", "Cofia bien puesta", 
                   "Mascarilla bien colocada", "Uso de protector auditivo", "Uñas cortas", 
                   "Guantes limpios", "Pestañas", "Barba/Bigote", "Medicamento autorizado", "Supervisor", "Observaciones"]
        ws.append(headers)

        # Agregar datos
        for record in data:
            ws.append(list(record))

        cur.close()
        conn.close()

        # Guardar el archivo en un objeto BytesIO y enviarlo
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="inspecciones.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

#PAGINA DE GESTIÓN DE PERSONAL

@app.route('/get-personnel', methods=['GET']) # Obtener la lista del personal
def get_personnel():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM personnel')
        data = cur.fetchall()

        cur.close()
        conn.close()

        return jsonify({"personnel": data}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get-areas', methods=['GET']) # Listar areas de la planta
def get_areas():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM area')
        data = cur.fetchall()

        cur.close()
        conn.close()

        return jsonify({"areas": data}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get-roles', methods=['GET'])# Listar cargos del personal
def get_roles():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT DISTINCT role FROM personnel')
        data = cur.fetchall()

        cur.close()
        conn.close()

        roles = [role[0] for role in data]

        return jsonify({"roles": roles}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/add-personnel', methods=['POST']) # Añadir una persona al personal
def add_personnel():
    try:
        data = request.json
        print(f"Datos recibidos para añadir personal: {data}")  # Ver los datos recibidos

        # Verifica que los datos que recibes son correctos
        if not all(key in data for key in ('name', 'role', 'id_area')):
            return jsonify({"error": "Faltan campos obligatorios."}), 400

        conn = get_db_connection()
        cur = conn.cursor()

        # Añadir los datos a la base de datos
        cur.execute('''
            INSERT INTO personnel (name, role, id_area)
            VALUES (%s, %s, %s)
        ''', (data['name'], data['role'], data['id_area']))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"message": "Personal agregado con éxito."}), 200

    except Exception as e:
        print(f"Error al añadir personal: {str(e)}")  # Mostrar el error
        return jsonify({"error": str(e)}), 500

@app.route('/update-personnel/<int:id>', methods=['PUT']) # Editar informacion de una persona 
def update_personnel(id):
    try:
        data = request.json
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(''' 
            UPDATE personnel
            SET name = %s, role = %s, id_area = %s
            WHERE id = %s
        ''', (data['name'], data['role'], data['id_area'], id))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"message": "Personal actualizado con éxito."}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/delete-personnel/<id>', methods=['DELETE']) # Eliminar una persona del personal
def delete_personnel(id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute('DELETE FROM personnel WHERE id = %s', (id,))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"message": "Personal eliminado con éxito."}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


#PAGINA FRECUENCIA DEL PERSONAL

@app.route('/inspection-frequency', methods=['GET'])
def get_inspection_frequency():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute('SELECT * FROM get_inspection_frequency()')
        data = cur.fetchall()

        # Ahora se mapea con la nueva columna 'ultima_inspeccion'
        results = [
            {
                'nombre_operario': row[0],
                'frecuencia': row[1],
                'ultima_inspeccion': row[2].strftime('%Y-%m-%d') if row[2] else None
            }
            for row in data
        ]

        cur.close()
        conn.close()

        return jsonify(results), 200

    except Exception as e:
        return jsonify({'error': f"Error interno del servidor: {str(e)}"}), 500



#PAGINA DE REVISION DE INSPECCION DEL PERSONAL

@app.route('/inspection-register', methods=['GET'])
def inspection_register():
    try:
        print("🔍 Entrando al endpoint /inspection-register")
        connection = get_db_connection()
        cursor = connection.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT * FROM inspection;")
        products = cursor.fetchall()
        cursor.close()
        connection.close()
        print("✅ Consulta realizada correctamente")
        return jsonify(products)
    except Exception as e:
        print(f"❌ Error en /inspection-register: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/inspection-register/<int:id>', methods=['PUT']) # Editar registro | Ingreso de material de empaque
def update_inspection_register(id):
    data = request.json
    connection = get_db_connection()
    cursor = connection.cursor()
    
    # Generar los campos dinámicamente
    update_fields = ", ".join([f"{key} = %s" for key in data.keys()])
    values = list(data.values()) + [id]
    
    query = f"UPDATE inspection SET {update_fields} WHERE id = %s;"
    cursor.execute(query, values)
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({"message": "Registro actualizado exitosamente."})

@app.route('/inspection-register/<int:id>', methods=['DELETE']) # Eliminar registro | Ingreso de material de empaque
def delete_inspection_register(id):
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("DELETE FROM inspection WHERE id = %s;", (id,))
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({"message": "Registro eliminado exitosamente."})

#PAGINA DE LOGIN 

@app.route('/login', methods=['POST']) # Comparar informacion para el inicio de sesión
def login():
    data = request.get_json()  # Recibe los datos del formulario (username, password)
    username = data.get('username')
    password = data.get('password')

    # Conectar a la base de datos y verificar el usuario y contraseña
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT * FROM users WHERE username = %s", (username,))
    user = cur.fetchone()
    cur.close()
    conn.close()

    if user and user['password'] == password:  # Verificación simple de la contraseña
        return jsonify({
            'success': True,
            'user_id': user['id']  # Suponiendo que cada usuario tiene un ID único
        })
    else:
        return jsonify({
            'success': False,
            'message': 'Usuario o contraseña incorrectos'
        }), 401

@app.route('/login-supervisor', methods=['POST']) # Comparar informacion para el inicio de sesión (Teniendo en cuenta el area)
def login_supervisor():
    data = request.get_json()  # Recibe los datos del formulario (username, password, area)
    username = data.get('username')
    password = data.get('password')
    user_area = data.get('area')  # El área fija de la página, enviada automáticamente desde el frontend

    # Conectar a la base de datos y verificar el usuario
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT * FROM users WHERE username = %s", (username,))
    user = cur.fetchone()
    cur.close()
    conn.close()

    if not user:
        return jsonify({
            'success': False,
            'message': 'Usuario no encontrado'
        }), 401

    # Verificar contraseña y área
    if user['password'] == password and user['area'] == user_area:
        return jsonify({
            'success': True,
            'user_id': user['id'],  # Suponiendo que cada usuario tiene un ID único
            'message': f"Bienvenido, {username}",  # Genera el token para la sesión
            'area': user['area']
        })
    else:
        return jsonify({
            'success': False,
            'message': 'Usuario, contraseña o área incorrectos'
        }), 401

#PAGINA "FORMULARIO DE ADITIVOS"

@app.route('/submit-additive-form', methods=['POST']) # Subir Formulario de Ingreso de Aditivos 
def submit_additive_form():
    try:
        # Obtiene los datos JSON
        data = request.json
        # Establecer conexión con la base de datos
        conn = get_db_connection()
        cur = conn.cursor()

        last_fumigation_date = data.get('last_fumigation_date') or None

        # Insertar los datos en la tabla correspondiente
        cur.execute('''
            INSERT INTO product_entry
            (entry_date, supplier, driver_name, driver_id, food_transport_permission,
             food_transport_validity, fumigation_record, last_fumigation_date, invoice_number,
             strange_smells, pests_evidence, clean_truck, uniformed_personnel, 
             floor_walls_roof_condition, truck_box_holes, disinfection_sticker,
             foreign_bodies, product, lot_number, shelf_life_check, 
             allergen_statement, graphic_system, product_accepted, rejection_reasons, 
             received_by, manufacture_date, expiry_date, package_quantity, total_weight,
             invoice_file_confirmation, truck_condition_image_confirmation, truck_plate_image_confirmation,
             technical_file_confirmation, liberation)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'EN ESPERA')
        ''', (
            data['entry_date'], data['supplier'], data['driver_name'], data['driver_id'],
            data['food_transport_permission'], data['food_transport_validity'],
            data['fumigation_record'], last_fumigation_date, data['invoice_number'],
            data['strange_smells'], data['pests_evidence'], data['clean_truck'],
            data['uniformed_personnel'], data['floor_walls_roof_condition'], 
            data['truck_box_holes'], data['disinfection_sticker'], data['foreign_bodies'], 
            data['product'], data['lot_number'], data['shelf_life_check'], 
            data['allergen_statement'], data['graphic_system'], data['product_accepted'], 
            data['rejection_reasons'], data['received_by'], data['manufacture_date'], 
            data['expiry_date'], data['package_quantity'], data['total_weight'], data['invoice_file_confirmation'],
            data['truck_condition_image_confirmation'], data['truck_plate_image_confirmation'], data['technical_file_confirmation']
        ))

        # Confirmar los cambios
        conn.commit()

        # Cerrar la conexión
        cur.close()
        conn.close()

        return jsonify({"message": "Formulario guardado correctamente"}), 200

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/submit-files', methods=['POST']) # Subir Archivos del Formulario de Ingreso de Aditivos
def submit_files():
    # Verificar si se recibió el campo 'supplier'
    if 'supplier' not in request.form:
        return jsonify({"error": "El campo 'supplier' es obligatorio."}), 400

    # Obtener el nombre del proveedor
    supplier_name = request.form.get('supplier', 'UNKNOWN').replace(' ', '_')

    product_name = request.form.get('product', 'UNKNOWN').replace(' ', '_')

    # Obtener la fecha actual para los nombres de los archivos
    current_date = datetime.now().strftime("%d-%m-%Y")

    # Definir carpetas específicas
    transporte_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'Transporte')
    producto_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'Producto')

    # Crear las carpetas si no existen
    os.makedirs(transporte_folder, exist_ok=True)
    os.makedirs(producto_folder, exist_ok=True)

    # Función para guardar el archivo con el nuevo nombre
    def save_file(file, folder, extra):
        if file:
            original_extension = os.path.splitext(file.filename)[1]
            new_filename = f"{current_date}_{supplier_name}_{product_name}_{extra}{original_extension}"
            file_path = os.path.join(folder, new_filename)
            file.save(file_path)
            return new_filename
        return None

    uploaded_files = {}

    # Guardar los archivos y asignarles un extra en su nombre
    if 'invoice_file' in request.files:
        uploaded_files['invoice_file'] = save_file(request.files['invoice_file'], transporte_folder, 'factura_guia')

    if 'truck_condition_image' in request.files:
        uploaded_files['truck_condition_image'] = save_file(request.files['truck_condition_image'], transporte_folder, 'estado_camion')

    if 'truck_plate_image' in request.files:
        uploaded_files['truck_plate_image'] = save_file(request.files['truck_plate_image'], transporte_folder, 'placa_del_camion')

    if 'technical_file' in request.files:
        uploaded_files['technical_file'] = save_file(request.files['technical_file'], producto_folder, 'ficha_certificado')

    return jsonify({"message": "Archivos subidos exitosamente", "files": uploaded_files}), 200

#PAGINA "ARCHIVOS | ADITIVOS"

@app.route('/get-files', methods=['GET']) # Listar los archivos |ADITIVOS
def get_files():
    try:
        # Verificar si las carpetas existen
        transporte_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'Transporte')
        producto_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'Producto')

        # Comprobar si las carpetas existen
        if not os.path.exists(transporte_folder):
            raise FileNotFoundError(f"La carpeta {transporte_folder} no existe.")
        if not os.path.exists(producto_folder):
            raise FileNotFoundError(f"La carpeta {producto_folder} no existe.")
        
        # Obtener los archivos de las carpetas
        transporte_files = os.listdir(transporte_folder)
        producto_files = os.listdir(producto_folder)

        # Si las carpetas están vacías, puedes devolver un mensaje
        if not transporte_files and not producto_files:
            return jsonify({"message": "No hay archivos en las carpetas."}), 200

        # Retornar los archivos como respuesta
        return jsonify({
            'transporte_files': transporte_files,
            'producto_files': producto_files
        }), 200

    except FileNotFoundError as e:
        print(f"Error: {e}")
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        print(f"Error inesperado: {e}")
        return jsonify({"error": "Error interno del servidor."}), 500

@app.route('/submit-just-one-file', methods=['POST']) # Subir un solo archivo | ADITIVOS
def submit_just_one_file():
    # Verificar que los campos necesarios estén presentes
    if 'supplier' not in request.form or 'fileType' not in request.form or 'date' not in request.form:
        return jsonify({"error": "Los campos 'supplier', 'fileType' y 'date' son obligatorios."}), 400

    supplier_name = request.form['supplier'].replace(' ', '_')
    product_name = request.form['product'].replace(' ', '_')
    file_type = request.form['fileType']
    # Validar y utilizar la fecha enviada
    try:
        selected_date = datetime.strptime(request.form['date'], "%d-%m-%Y").strftime("%d-%m-%Y")
    except ValueError:
        return jsonify({"error": "El formato de la fecha es inválido. Usa 'dd-mm-yyyy'."}), 400

    # Determinar la ruta de la carpeta donde se guardará el archivo
    if file_type in ['factura_guia', 'estado_camion', 'placa_camion']:
        folder_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Transporte')
    elif file_type == 'ficha_certificado':
        folder_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Producto')
    else:
        return jsonify({"error": "Tipo de archivo no válido."}), 400

    os.makedirs(folder_path, exist_ok=True)

    # Guardar el archivo
    file = request.files.get('file')
    if file:
        original_extension = os.path.splitext(file.filename)[1]
        new_filename = f"{selected_date}_{supplier_name}_{product_name}_{file_type}{original_extension}"
        file_path = os.path.join(folder_path, new_filename)

        print(f"Guardando archivo en: {file_path}")
        file.save(file_path)

        return jsonify({"message": "Archivo subido exitosamente", "filename": new_filename}), 200
    else:
        return jsonify({"error": "No se ha enviado ningún archivo."}), 400

@app.route('/download/<folder>/<filename>', methods=['GET']) # Descargar archivo elegido | ADITIVOS
def download_file(folder, filename):
    # Construir la ruta completa
    folder_path = f"{UPLOAD_FOLDER}/{folder}"
    try:
        # Enviar el archivo desde la carpeta correspondiente
        return send_from_directory(folder_path, filename, as_attachment=True)
    except FileNotFoundError:
        # Manejo de error si el archivo no existe
        abort(404, description="Archivo no encontrado.")

@app.route('/delete-file', methods=['DELETE']) # Eliminar archivo elegido | ADITIVOS
def delete_file():
    try:
        # Obtener datos del archivo y la carpeta desde la solicitud
        data = request.get_json()
        file_name = data['file']
        folder = data['folder']

        # Determinar la ruta del archivo a eliminar
        if folder not in ['Transporte', 'Producto']:
            return jsonify({'error': 'Carpeta no válida'}), 400

        file_path = os.path.join(UPLOAD_FOLDER, folder, file_name)

        # Verificar si el archivo existe
        if os.path.exists(file_path):
            os.remove(file_path)
            return jsonify({'message': 'Archivo eliminado correctamente'}), 200
        else:
            return jsonify({'error': 'Archivo no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# PÁGINA "REGISTRO DE ADITIVOS"

@app.route('/products', methods=['GET']) # Listar registro del ingreso de aditivos
def get_products():
    connection = get_db_connection()
    cursor = connection.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT * FROM product_entry;")
    products = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify(products)

@app.route('/products/<int:product_id>', methods=['PUT']) # Editar un registo | Ingreso de Aditivos
def update_product(product_id):
    data = request.json
    connection = get_db_connection()
    cursor = connection.cursor()
    
    # Generar los campos dinámicamente
    update_fields = ", ".join([f"{key} = %s" for key in data.keys()])
    values = list(data.values()) + [product_id]
    
    query = f"UPDATE product_entry SET {update_fields} WHERE id = %s;"
    cursor.execute(query, values)
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({"message": "Producto actualizado exitosamente."})

@app.route('/products/<int:product_id>', methods=['DELETE']) # Eliminar un registro | Ingreso de Aditivos
def delete_product(product_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("DELETE FROM product_entry WHERE id = %s;", (product_id,))
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({"message": "Producto eliminado exitosamente."})

@app.route('/download-product-table', methods=['GET']) # Descargar registro de ingreso de Aditivos
def download_product_table():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Consulta para seleccionar todas las columnas de la tabla
        cur.execute('''
            SELECT id, entry_date, supplier, driver_name, driver_id, 
                   food_transport_permission, food_transport_validity, 
                   fumigation_record, last_fumigation_date, invoice_number, 
                   strange_smells, pests_evidence, clean_truck, uniformed_personnel, 
                   floor_walls_roof_condition, truck_box_holes, disinfection_sticker, 
                   foreign_bodies, product, lot_number, 
                   package_quantity, total_weight, manufacture_date, expiry_date, 
                   shelf_life_check, allergen_statement, graphic_system, 
                   product_accepted, rejection_reasons, received_by, 
                   invoice_file_confirmation, truck_condition_image_confirmation, 
                   truck_plate_image_confirmation, technical_file_confirmation, liberation
            FROM product_entry
            ORDER BY entry_date DESC
        ''')
        data = cur.fetchall()

        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        # Encabezados en español
        headers = [
            "ID", "Fecha de Ingreso", "Proveedor", "Nombre del Conductor", "ID del Conductor",
            "Permiso de Transporte de Alimentos", "Vigencia del Permiso", "Registro de Fumigación",
            "Última Fecha de Fumigación", "Número de Factura", "Olores Extraños", "Evidencia de Plagas",
            "Camión Limpio", "Personal Uniformado", "Estado del Piso, Paredes y Techo",
            "Huecos en la Caja del Camión", "Etiqueta de Desinfección", "Cuerpos Extraños",
            "Producto", "Número de Lote", "Cantidad de Paquetes", "Peso Total",
            "Fecha de Fabricación", "Fecha de Expiración", "Verificación de Vida Útil",
            "Declaración de Alérgenos", "Sistema Gráfico", "Producto Aceptado",
            "Razones de Rechazo", "Recibido Por", "Confirmación de Factura",
            "Confirmación de Imagen del Estado del Camión",
            "Confirmación de Imagen de la Placa del Camión", "Confirmación de Archivo Técnico", "Liberación"
        ]
        ws.append(headers)

        # Agregar datos
        for record in data:
            ws.append(list(record))

        cur.close()
        conn.close()

        # Guardar el archivo en un objeto BytesIO y enviarlo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="registro-aditivos.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/provider', methods=['GET']) 
def get_provider():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT provider_name FROM providers ORDER BY id ASC")  # Ajusta el nombre de la tabla y columna según tu base de datos
        providers = [row['provider_name'] for row in cursor.fetchall()]
        return jsonify(providers)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/product', methods=['GET'])
def get_product():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT product_name FROM products WHERE product_type IN ('ADITIVO', 'MATERIA PRIMA') ORDER BY id ASC")
        products = [row['product_name'] for row in cursor.fetchall()]
        return jsonify(products)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

#Pagina "Agregar Productos y Proveedores | ADITIVOS"

@app.route('/get-providers', methods=['GET']) # Listar proveedores | Aditivos
def get_providers_list():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT id, provider_name FROM providers")
        providers = cursor.fetchall()
        return jsonify(providers)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/get-products', methods=['GET']) # Listar productos | Aditivos
def get_products_list():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT p.id, p.product_name, p.product_type, pr.provider_name
            FROM products p
            JOIN providers pr ON p.provider_id = pr.id
        """)
        products = cursor.fetchall()
        return jsonify(products)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/add-provider', methods=['POST']) # Añadir proveedores | Aditivos
def add_provider():
    data = request.get_json()
    provider_name = data.get('provider_name').upper()  # Convertir a mayúsculas

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO providers (provider_name) VALUES (%s) RETURNING id", (provider_name,))
        new_provider_id = cursor.fetchone()[0]
        conn.commit()
        return jsonify({'id': new_provider_id, 'provider_name': provider_name}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/add-product', methods=['POST']) # Añadir producto | Aditivos
def add_product():
    data = request.get_json()
    product_name = data.get('product_name').upper()  # Convertir a mayúsculas
    provider_id = data.get('provider_id')
    product_type = data.get('product_type')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO products (product_name, provider_id, product_type) VALUES (%s, %s, %s) RETURNING id", 
                       (product_name, provider_id, product_type))
        new_product_id = cursor.fetchone()[0]
        conn.commit()
        return jsonify({'id': new_product_id, 'product_name': product_name, 'provider_id': provider_id, 'product_type': product_type}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/delete-provider/<int:id>', methods=['DELETE']) # Eliminar proveedor | Aditivos
def delete_provider(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM providers WHERE id = %s RETURNING id", (id,))
        deleted_provider = cursor.fetchone()

        if deleted_provider:
            conn.commit()
            return jsonify({'message': 'Proveedor eliminado exitosamente'}), 200
        else:
            return jsonify({'error': 'Proveedor no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/delete-product/<int:id>', methods=['DELETE']) # Eliminar producto | Aditivos
def delete_products(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM products WHERE id = %s RETURNING id", (id,))
        deleted_product = cursor.fetchone()

        if deleted_product:
            conn.commit()
            return jsonify({'message': 'Producto eliminado exitosamente'}), 200
        else:
            return jsonify({'error': 'Producto no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/update-provider/<int:id>', methods=['PUT']) # Editar proveedor seleccionado | Aditivos
def update_provider(id):
    data = request.get_json()
    new_provider_name = data.get('provider_name')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE providers SET provider_name = %s WHERE id = %s RETURNING id, provider_name", 
                       (new_provider_name, id))
        updated_provider = cursor.fetchone()

        if updated_provider:
            conn.commit()
            return jsonify({'id': updated_provider[0], 'provider_name': updated_provider[1]}), 200
        else:
            return jsonify({'error': 'Proveedor no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/update-product/<int:id>', methods=['PUT']) # Editar producto seleccionado | Aditivos
def update_products(id):
    data = request.get_json()
    new_product_name = data.get('product_name')
    new_provider_id = data.get('provider_id')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            UPDATE products
            SET product_name = %s, provider_id = %s
            WHERE id = %s
            RETURNING id, product_name, provider_id
            """,
            (new_product_name, new_provider_id, id),
        )
        updated_product = cursor.fetchone()

        if updated_product:
            conn.commit()
            return jsonify({
                'id': updated_product[0],
                'product_name': updated_product[1],
                'provider_id': updated_product[2]
            }), 200
        else:
            return jsonify({'error': 'Producto no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

#Pagina "Liberacion de producto"

@app.route('/get-pending-products', methods=['GET']) #Listar Productos por Liberar | Aditivos
def get_pending_products():
    try:
        # Establecer conexión con la base de datos
        conn = get_db_connection()
        cur = conn.cursor()

        # Consultar productos con estado "EN ESPERA"
        cur.execute('''
            SELECT id, entry_date, product, supplier
            FROM product_entry
            WHERE liberation = 'EN ESPERA'
        ''')

        # Obtener los resultados
        products = cur.fetchall()

        # Convertir los resultados en una lista de diccionarios
        product_list = [
            {
                "id": row[0],
                "entry_date": row[1].strftime('%Y-%m-%d'),  # Convertir a formato de fecha
                "product": row[2],
                "supplier": row[3]
            }
            for row in products
        ]

        # Cerrar conexión
        cur.close()
        conn.close()

        return jsonify(product_list), 200
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/submit-release', methods=['POST']) # Liberar el producto seleccionado | Aditivos
def submit_release():
    try:
        # Obtener los datos enviados por el cliente
        data = request.json
        product_id = data.get('product_id')
        analysis_match = data.get('analysis_match')
        release_criteria = data.get('release_criteria')
        releaser = data.get('releaser')

        if product_id is None or analysis_match is None or release_criteria is None:
            return jsonify({"error": "Faltan datos obligatorios"}), 400

        # Determinar el estado de liberación basado en las respuestas
        if (analysis_match == 'SI' and release_criteria == 'SI') or (analysis_match == 'NO APLICA' and release_criteria == 'SI'):
            release_status = 'SI'
        else:
            release_status = 'NO'

        # Conectar a la base de datos
        conn = get_db_connection()
        cur = conn.cursor()

        # Insertar las respuestas en la tabla de liberaciones
        cur.execute('''
            INSERT INTO additive_release (product_id, analysis_match, release_criteria, release_status, releaser)
            VALUES (%s, %s, %s, %s, %s)
        ''', (product_id, analysis_match, release_criteria, release_status, releaser))

        # Actualizar el estado de liberación en la tabla product_entry
        cur.execute('''
            UPDATE product_entry 
            SET liberation = %s 
            WHERE id = %s
        ''', (release_status, product_id))

        # Confirmar los cambios
        conn.commit()

        # Cerrar la conexión
        cur.close()
        conn.close()

        return jsonify({"message": "Liberación registrada correctamente"}), 200

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/get-additive-releases', methods=['GET'])
def get_additive_releases():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT ar.id,
               p.product_name AS name_product,
               ar.release_date,
               ar.releaser,
               ar.release_criteria,
               ar.analysis_match,
               ar.release_status
        FROM additive_release ar
        JOIN product_entry pe ON ar.product_id = pe.id
        JOIN products p ON pe.product = p.product_name
        ORDER BY ar.release_date DESC
    """)
    rows = cur.fetchall()
    releases = [
        {
            "id": r[0],
            "product_name": r[1],
            "release_date": r[2],
            "releaser": r[3],
            "release_criteria": r[4],
            "analysis_match": r[5],
            "release_status": r[6],
        }
        for r in rows
    ]
    cur.close()
    conn.close()
    return jsonify(releases)



#Pagina "Añadir Proveedores o Material de Empaque"

@app.route('/add-brand', methods=['POST']) # Añadir proveedores | Aditivos
def add_brand():
    data = request.get_json()
    brand_name = data.get('brand_name').upper()  # Convertir a mayúsculas

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO brand (brand_name) VALUES (%s) RETURNING id", (brand_name,))
        new_brand_id = cursor.fetchone()[0]
        conn.commit()
        return jsonify({'id': new_brand_id, 'provider_name': brand_name}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/get-providers-material', methods=['GET']) # Listar proveedores | Material de empaque
def get_providers_material_list():

    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT id, provider_name FROM providers_material ORDER BY provider_name ASC")
        providers = cursor.fetchall()
        return jsonify(providers)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/get-brand-material', methods=['GET']) # Obtener los nombres de las marcas | Material de Empaque
def get_brand_material():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT id, brand_name FROM brand ORDER BY brand_name ASC")
        providers = cursor.fetchall()
        return jsonify(providers)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/get-materials', methods=['GET']) # Listar Material de Empaque
def get_products_material_list():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT m.id, m.code, m.code_sap, m.material_name, m.material_type, prm.provider_name, b.brand_name
            FROM materials m
            JOIN providers_material prm ON m.provider_id = prm.id
            JOIN brand b ON m.brand_id = b.id; 

        """)
        materials = cursor.fetchall()
        return jsonify(materials)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/add-provider-material', methods=['POST']) # Añadir proveedor | Material de Empaque 
def add_provider_material():
    data = request.get_json()
    provider_name = data.get('provider_name').upper()  # Convertir a mayúsculas

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO providers_material (provider_name) VALUES (%s) RETURNING id", (provider_name,))
        new_provider_id = cursor.fetchone()[0]
        conn.commit()
        return jsonify({'id': new_provider_id, 'provider_name': provider_name}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/add-material', methods=['POST']) # Añadir nuevo material de empaque
def add_material():
    data = request.get_json()
    material_name = data.get('material_name').upper()  # Convertir a mayúsculas
    provider_id = data.get('provider_id')
    code = data.get('code')
    code_sap = data.get('code_sap')
    material_type = data.get('material_type').upper()
    brand_id = data.get('brand_id')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO materials (code, code_sap, material_name, provider_id, material_type, brand_id) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id", 
                       (code, code_sap, material_name, provider_id, material_type, brand_id))
        new_material_id = cursor.fetchone()[0]
        conn.commit()
        return jsonify({'message': 'Material guardado exitosamente','id': new_material_id,'code': code, 'code_sap': code_sap, 'product_name': material_name, 'provider_id': provider_id, 'material_type': material_type, 'brand_id': brand_id}), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/delete-provider-material/<int:id>', methods=['DELETE']) # Eliminar proveedot | Material de empaque
def delete_provider_material(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM providers_material WHERE id = %s RETURNING id", (id,))
        deleted_provider_material = cursor.fetchone()

        if deleted_provider_material:
            conn.commit()
            return jsonify({'message': 'Proveedor eliminado exitosamente'}), 200
        else:
            return jsonify({'error': 'Proveedor no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/delete-material/<int:id>', methods=['DELETE']) # Eliminar material seleccionado
def delete_material(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM materials WHERE id = %s RETURNING id", (id,))
        deleted_product = cursor.fetchone()

        if deleted_product:
            conn.commit()
            return jsonify({'message': 'Producto eliminado exitosamente'}), 200
        else:
            return jsonify({'error': 'Producto no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/update-provider-material/<int:id>', methods=['PUT']) # Editar proveedor | Material de empaque
def update_provider_material(id):
    data = request.get_json()
    new_provider_name = data.get('provider_name')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE providers_material SET provider_name = %s WHERE id = %s RETURNING id, provider_name", 
                       (new_provider_name, id))
        updated_provider = cursor.fetchone()

        if updated_provider:
            conn.commit()
            return jsonify({'id': updated_provider[0], 'provider_name': updated_provider[1]}), 200
        else:
            return jsonify({'error': 'Proveedor no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/update-material/<int:id>', methods=['PUT']) # Editar material de empaque seleccionado
def update_materials(id):
    data = request.get_json()
    new_material_name = data.get('material_name')
    new_provider_id = data.get('provider_id')
    new_code = data.get('code')
    new_code_sap = data.get('code_sap')
    new_material_type = data.get('material_type')
    new_brand_id = data.get('brand_id')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            UPDATE materials
            SET code = %s, code_sap = %s, material_name = %s, provider_id = %s, material_type = %s, brand_id = %s
            WHERE id = %s
            RETURNING id, code, code_sap, material_name, provider_id, material_type, brand_id
            """,
            (new_code, new_code_sap, new_material_name, new_provider_id, new_material_type, new_brand_id, id),
        )
        updated_material = cursor.fetchone()

        if updated_material:
            conn.commit()
            return jsonify({
                'id': updated_material[0],
                'code': updated_material[1],
                'code_sap': updated_material[2],
                'material_name': updated_material[3],
                'provider_id': updated_material[4],
                'material_type': updated_material[5],
                'brand_id': updated_material[6]               
            }), 200
        else:
            return jsonify({'error': 'Producto no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

#Página "Formulario de ingreso Material de Empaque"

@app.route('/submit-material-form', methods=['POST']) # Subir formulario material de empaque
def submit_material_form():
    try:
        # Obtiene los datos JSON
        data = request.json
        # Establecer conexión con la base de datos
        conn = get_db_connection()
        cur = conn.cursor()

        # Insertar los datos en la tabla correspondiente
        cur.execute('''
            INSERT INTO material_entry
            (entry_date, supplier, driver_name, driver_id, invoice_number,
             strange_smells, pests_evidence, clean_truck, uniformed_personnel, 
             floor_walls_roof_condition, truck_box_holes,
             foreign_bodies, brand, product, lot_number, shelf_life_check, 
             allergen_statement, product_accepted, rejection_reasons, 
             received_by, manufacture_date, package_quantity, unit_quantity, total_weight,
             invoice_file_confirmation, truck_condition_image_confirmation, truck_plate_image_confirmation,
             technical_file_confirmation)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s )
        ''', (
            data['entry_date'], data['supplier'], data['driver_name'], data['driver_id'],
            data['invoice_number'],
            data['strange_smells'], data['pests_evidence'], data['clean_truck'],
            data['uniformed_personnel'], data['floor_walls_roof_condition'], 
            data['truck_box_holes'], data['foreign_bodies'], data['brand'],
            data['product'], data['lot_number'], data['shelf_life_check'], 
            data['allergen_statement'], data['product_accepted'], 
            data['rejection_reasons'], data['received_by'], data['manufacture_date'], 
            data['package_quantity'], data['unit_quantity'], data['total_weight'], data['invoice_file_confirmation'],
            data['truck_condition_image_confirmation'], data['truck_plate_image_confirmation'], data['technical_file_confirmation']
            
        ))

        # Confirmar los cambios
        conn.commit()

        # Cerrar la conexión
        cur.close()
        conn.close()

        return jsonify({"message": "Formulario guardado correctamente"}), 200

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/submit-materials-files', methods=['POST']) # Subir archivos | Material de empaque
def submit_material_files():
    # Verificar si se recibió el campo 'supplier'
    if 'supplier' not in request.form:
        return jsonify({"error": "El campo 'supplier' es obligatorio."}), 400

    # Obtener el nombre del proveedor
    supplier_name = request.form.get('supplier', 'UNKNOWN').replace(' ', '_')

    # Obtener el nombre del producto
    product_name = request.form.get('product', 'UNKNOWN').replace(' ', '_')

    # Obtener el nombre de la marca (agregar este paso)
    brand_name = request.form.get('brand', 'UNKNOWN').replace(' ', '_')  # Asegúrate de que este campo esté presente en el formulario

    provided_date = request.form.get('entry_date', None)
    if provided_date:
        # Validar el formato de la fecha (YYYY-MM-DD)
        try:
            formatted_date = datetime.strptime(provided_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        except ValueError:
            return jsonify({"error": "El formato de la fecha es inválido. Use 'YYYY-MM-DD'."}), 400
    else:
        formatted_date = datetime.now().strftime("%d-%m-%Y")
    

    # Definir carpetas específicas
    transporte_folder = os.path.join(app.config['UPLOAD_MATERIAL_FOLDER'], 'Transporte')
    producto_folder = os.path.join(app.config['UPLOAD_MATERIAL_FOLDER'], 'Producto')

    # Crear las carpetas si no existen
    os.makedirs(transporte_folder, exist_ok=True)
    os.makedirs(producto_folder, exist_ok=True)

    # Función para guardar el archivo con el nuevo nombre
    def save_file(file, folder, extra):
        if file:
            original_extension = os.path.splitext(file.filename)[1]
            new_filename = f"{formatted_date}_{supplier_name}_{brand_name}_{product_name}_{extra}{original_extension}"  # Incluir la marca aquí
            file_path = os.path.join(folder, new_filename)
            file.save(file_path)
            return new_filename
        return None

    uploaded_files = {}

    # Guardar los archivos y asignarles un extra en su nombre
    if 'invoice_file' in request.files:
        uploaded_files['invoice_file'] = save_file(request.files['invoice_file'], transporte_folder, 'factura_guia')

    if 'truck_condition_image' in request.files:
        uploaded_files['truck_condition_image'] = save_file(request.files['truck_condition_image'], transporte_folder, 'estado_camion')

    if 'truck_plate_image' in request.files:
        uploaded_files['truck_plate_image'] = save_file(request.files['truck_plate_image'], transporte_folder, 'placa_del_camion')

    if 'technical_file' in request.files:
        uploaded_files['technical_file'] = save_file(request.files['technical_file'], producto_folder, 'ficha_certificado')

    return jsonify({"message": "Archivos subidos exitosamente", "files": uploaded_files}), 200

@app.route('/providers-material', methods=['GET']) # Listar proveedores | Material de empaque
def get_providers_material():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT provider_name FROM providers_material ORDER BY provider_name ASC")  # Ajusta el nombre de la tabla y columna según tu base de datos
        providers = [row['provider_name'] for row in cursor.fetchall()]
        return jsonify(providers)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/brand', methods=['GET']) # Listar marcas | Material de empaque
def get_brand():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT brand_name FROM brand ORDER BY brand_name ASC")
        brands = [row['brand_name'] for row in cursor.fetchall()]
        return jsonify(brands)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/get-code', methods=['GET']) # Listar codigos de cada material
def get_code():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT code FROM materials")
        codes = [row['code'] for row in cursor.fetchall()]
        return jsonify(codes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/get-product-type/<code>', methods=['GET']) # Mostrar el tipo de material de empaque elegido
def get_product_type_by_code(code):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT material_type FROM materials WHERE code = %s", (code,))
        result = cursor.fetchone()
        if result:
            return jsonify(result)  # Devuelve {"material_name": "Nombre del producto"}
        else:
            return jsonify({'error': 'Código no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/get-product-name/<code>', methods=['GET']) # Mostrar el nombre del material de empaque elegido
def get_product_name_by_code(code):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT material_name FROM materials WHERE code = %s", (code,))
        result = cursor.fetchone()
        if result:
            return jsonify(result)  # Devuelve {"material_name": "Nombre del producto"}
        else:
            return jsonify({'error': 'Código no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/get-products-by-brand/<brand>', methods=['GET']) 
def get_products_by_brand(brand):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT m.code
            FROM materials m
            JOIN brand b ON m.brand_id = b.id
            WHERE b.brand_name = %s
        """, (brand,))
        products = cursor.fetchall()
        return jsonify(products)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/get-products-by-brand-provider/<brand>/<provider>', methods=['GET']) # Listar materiales de acuerdo a la marca y proveedor elegidos
def get_products_by_brand_provider(brand, provider):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        query = """
            SELECT p.code
            FROM materials p
            INNER JOIN brand b ON p.brand_id = b.id
            INNER JOIN providers_material pr ON p.provider_id = pr.id
            WHERE b.brand_name = %s AND pr.provider_name = %s
        """
        params = [brand, provider]

        query += " ORDER BY p.code ASC"

        cursor.execute(query, params)
        products = cursor.fetchall()

        print("Productos devueltos:", products)

        return jsonify(products)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()


#"Pagina de archivos Material de empaque"

@app.route('/get-material-files', methods=['GET']) # Listar archivos | Material de empaque
def get_material_files():
    try:
        # Verificar si las carpetas existen
        transporte_folder = os.path.join(app.config['UPLOAD_MATERIAL_FOLDER'], 'Transporte')
        producto_folder = os.path.join(app.config['UPLOAD_MATERIAL_FOLDER'], 'Producto')

        # Comprobar si las carpetas existen
        if not os.path.exists(transporte_folder):
            raise FileNotFoundError(f"La carpeta {transporte_folder} no existe.")
        if not os.path.exists(producto_folder):
            raise FileNotFoundError(f"La carpeta {producto_folder} no existe.")
        
        # Obtener los archivos de las carpetas
        transporte_files = os.listdir(transporte_folder)
        producto_files = os.listdir(producto_folder)

        # Si las carpetas están vacías, puedes devolver un mensaje
        if not transporte_files and not producto_files:
            return jsonify({"message": "No hay archivos en las carpetas."}), 200

        # Retornar los archivos como respuesta
        return jsonify({
            'transporte_files': transporte_files,
            'producto_files': producto_files
        }), 200

    except FileNotFoundError as e:
        print(f"Error: {e}")
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        print(f"Error inesperado: {e}")
        return jsonify({"error": "Error interno del servidor."}), 500

@app.route('/submit-just-one-file-material', methods=['POST']) # Subir un archivo | Material de empaque
def submit_just_one_file_material():
    # Verificar que los campos necesarios estén presentes
    if 'supplier' not in request.form or 'fileType' not in request.form or 'date' not in request.form or 'brand' not in request.form:
        return jsonify({"error": "Los campos 'supplier', 'fileType' y 'date' son obligatorios."}), 400

    supplier_name = request.form['supplier'].replace(' ', '_')
    product_name = request.form['product'].replace(' ', '_')
    brand_name = request.form['brand'].replace(' ', '_')
    file_type = request.form['fileType']
    # Validar y utilizar la fecha enviada
    try:
        selected_date = datetime.strptime(request.form['date'], "%d-%m-%Y").strftime("%d-%m-%Y")
    except ValueError:
        return jsonify({"error": "El formato de la fecha es inválido. Usa 'dd-mm-yyyy'."}), 400

    # Determinar la ruta de la carpeta donde se guardará el archivo
    if file_type in ['estado_camion', 'placa_camion', 'factura_guia']:
        folder_path = os.path.join(app.config['UPLOAD_MATERIAL_FOLDER'], 'Transporte')
    elif file_type in ['ficha_certificado']:
        folder_path = os.path.join(app.config['UPLOAD_MATERIAL_FOLDER'], 'Producto')
    else:
        return jsonify({"error": "Tipo de archivo no válido."}), 400

    os.makedirs(folder_path, exist_ok=True)

    # Guardar el archivo
    file = request.files.get('file')
    if file:
        original_extension = os.path.splitext(file.filename)[1]
        new_filename = f"{selected_date}_{supplier_name}_{brand_name}_{product_name}_{file_type}{original_extension}"
        file_path = os.path.join(folder_path, new_filename)

        print(f"Guardando archivo en: {file_path}")
        file.save(file_path)

        return jsonify({"message": "Archivo subido exitosamente", "filename": new_filename}), 200
    else:
        return jsonify({"error": "No se ha enviado ningún archivo."}), 400

@app.route('/download-material-file/<folder>/<filename>', methods=['GET']) # Descargar archivo seleccionado | Material de empque
def download_material_file(folder, filename):
    # Construir la ruta completa
    folder_path = f"{UPLOAD_MATERIAL_FOLDER}/{folder}"
    try:
        # Enviar el archivo desde la carpeta correspondiente
        return send_from_directory(folder_path, filename, as_attachment=True)
    except FileNotFoundError:
        # Manejo de error si el archivo no existe
        abort(404, description="Archivo no encontrado.")

@app.route('/delete-material-file', methods=['DELETE']) # Eliminar archivo seleccionado | Material de empqeu
def delete_material_file():
    try:
        # Obtener datos del archivo y la carpeta desde la solicitud
        data = request.get_json()
        file_name = data['file']
        folder = data['folder']

        # Determinar la ruta del archivo a eliminar
        if folder not in ['Transporte', 'Producto']:
            return jsonify({'error': 'Carpeta no válida'}), 400

        file_path = os.path.join(UPLOAD_MATERIAL_FOLDER, folder, file_name)

        # Verificar si el archivo existe
        if os.path.exists(file_path):
            os.remove(file_path)
            return jsonify({'message': 'Archivo eliminado correctamente'}), 200
        else:
            return jsonify({'error': 'Archivo no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/provider-material', methods=['GET']) # Listar proveedores | Material de Empaque
def get_provider_material():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT provider_name FROM providers_material ORDER BY id ASC")  # Ajusta el nombre de la tabla y columna según tu base de datos
        providers = [row['provider_name'] for row in cursor.fetchall()]
        return jsonify(providers)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

@app.route('/material', methods=['GET']) 
def get_material():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT material_name FROM packaging_material ORDER BY material_name ASC")
        products = [row['material_name'] for row in cursor.fetchall()]
        return jsonify(products)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            conn.close()

#PAGINA "REGISTRO DE MATERIAL DE EMPAQUE"

@app.route('/materials', methods=['GET']) # Listar registro de ingreso de material de empaque
def get_material_list():
    connection = get_db_connection()
    cursor = connection.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT * FROM material_entry;")
    products = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify(products)

@app.route('/materials/<int:material_id>', methods=['PUT']) # Editar registro | Ingreso de material de empaque
def update_material_list(material_id):
    data = request.json
    connection = get_db_connection()
    cursor = connection.cursor()
    
    # Generar los campos dinámicamente
    update_fields = ", ".join([f"{key} = %s" for key in data.keys()])
    values = list(data.values()) + [material_id]
    
    query = f"UPDATE material_entry SET {update_fields} WHERE id = %s;"
    cursor.execute(query, values)
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({"message": "Producto actualizado exitosamente."})

@app.route('/materials/<int:material_id>', methods=['DELETE']) # Eliminar registro | Ingreso de material de empaque
def delete_material_list(material_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("DELETE FROM material_entry WHERE id = %s;", (material_id,))
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({"message": "Producto eliminado exitosamente."})

@app.route('/download-material-table', methods=['GET']) # Descargar registro | Ingreso de material de empaque
def download_material_table():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Consulta para seleccionar todas las columnas de la tabla
        cur.execute('''
            SELECT id, entry_date, supplier, driver_name, driver_id, invoice_number,
             strange_smells, pests_evidence, clean_truck, uniformed_personnel, 
             floor_walls_roof_condition, truck_box_holes,
             foreign_bodies, brand, product, lot_number, shelf_life_check, 
             allergen_statement, product_accepted, rejection_reasons, 
             received_by, manufacture_date, package_quantity, unit_quantity, total_weight,
             invoice_file_confirmation, truck_condition_image_confirmation, truck_plate_image_confirmation,
             technical_file_confirmation
            FROM material_entry
            ORDER BY entry_date DESC
        ''')
        data = cur.fetchall()

        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Material de Empaque"

        # Encabezados en español
        headers = [
            "ID", "Fecha de Ingreso", "Proveedor", "Nombre del Conductor", "ID del Conductor",
            "Número de Factura", "Olores Extraños", "Evidencia de Plagas",
            "Camión Limpio", "Personal Uniformado", "Estado del Piso, Paredes y Techo",
            "Huecos en la Caja del Camión", "Cuerpos Extraños", "Marca",
            "Producto", "Número de Lote", "Cantidad de Paquetes", "Cantidad de unidades", "Peso Total",
            "Fecha de Fabricación", "Verificación de Vida Útil",
            "Declaración de Alérgenos", "Producto Aceptado",
            "Razones de Rechazo", "Recibido Por", "Confirmación de Factura",
            "Confirmación de Imagen del Estado del Camión",
            "Confirmación de Imagen de la Placa del Camión", "Confirmación de Archivo Técnico"
        ]
        ws.append(headers)

        # Agregar datos
        for record in data:
            ws.append(list(record))

        cur.close()
        conn.close()

        # Guardar el archivo en un objeto BytesIO y enviarlo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="registro-material-de-empaque.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

#PÁGINA CONTROL DE PESOS

@app.route('/get-product-info', methods=['GET']) # Obtener datos del producto producto | Control del pesos
def get_product_info():
    # Obtener el EAN13 desde los parámetros de la solicitud
    ean13 = request.args.get('ean13')

    if not ean13:
        return jsonify({"error": "EAN13 es requerido"}), 400

    try:
        # Conectar a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Consulta para obtener la información del producto, uniendo la tabla 'article' con la tabla 'brand'
        query = """
        SELECT 
            a.ean13, 
            a.weight AS peso_neto, 
            a.format AS formato, 
            a.brand_id, 
            b.brand_name AS marca
        FROM articles a
        JOIN brand b ON a.brand_id = b.id
        WHERE a.ean13 = %s
        """
        cursor.execute(query, (ean13,))
        product = cursor.fetchone()

        # Cerrar la conexión
        cursor.close()
        conn.close()

        # Verificar si el producto existe
        if product:
            return jsonify(product), 200
        else:
            return jsonify({"error": "Producto no encontrado"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get-balers', methods=['GET']) # Obtener los nombres de las empacadoras | Control de pesos
def get_balers():
    try:
        # Conectar a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Consulta para obtener todas las empacadoras
        query = "SELECT id, baler_name FROM balers"
        cursor.execute(query)
        balers = cursor.fetchall()

        # Cerrar conexión
        cursor.close()
        conn.close()

        return jsonify(balers), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/save-weight-control', methods=['POST']) # Subir el nuevo registro de control de pesos
def save_weight_control():
    try:
        data = request.json

        # Extraer datos principales
        date = data.get('date', datetime.today().strftime('%Y-%m-%d'))
        baler = data.get('baler')
        net_weight = float(data.get('net_weight'))  # Peso nominal (Qnom)
        format_ = data.get('format')
        brand = data.get('brand')
        lot = data.get('lot')
        ean13 = data.get('ean13')
        manufacture_date = data.get('manufacture_date')
        expiry_date = data.get('expiry_date')

        # Extraer los pesos y asegurarse de que sean numéricos
        pesos = [float(p) for p in data.get('weights', []) if p not in [None, ""]]

        # Cálculo de tolerancia basado en la fórmula proporcionada
        if net_weight == 100:
            error_T1 = net_weight * 0.045  # 4.5% de net_weight
        elif net_weight >= 200 and net_weight < 300:
            error_T1 = 9
        elif net_weight >= 300 and net_weight < 500:
            error_T1 = net_weight * 0.03  # 3% de net_weight
        elif net_weight >= 500 and net_weight < 1000:
            error_T1 = 15
        elif net_weight >= 1000 and net_weight < 10000:
            error_T1 = net_weight * 0.015  # 1.5% de net_weight
        elif net_weight >= 10000 and net_weight < 15000:
            error_T1 = 150
        elif net_weight >= 15000 and net_weight < 50000:
            error_T1 = net_weight * 0.01  # 1% de net_weight
        else:
            error_T1 = "ERROR TOLERANCIA"  # Error en tolerancia si no cae en los rangos

        if error_T1 == "ERROR TOLERANCIA":
            return jsonify({"error": "Error en los rangos de tolerancia para el peso nominal"}), 400

        # Cálculo de ERROR T2
        error_T2 = error_T1 * 2

        # Cálculo de límites máximos operativos
        limite_maximo_operativo = round(net_weight + (net_weight * 0.02), 2)
        limite_minimo_operativo = round(net_weight - (net_weight * 0.03), 2)

        # Calcular estadísticas de peso
        if pesos:
            average = float(np.mean(pesos))  # Convertir np.float64 a float
            minimum = float(np.min(pesos))
            maximum = float(np.max(pesos))
            std_dev = float(np.std(pesos, ddof=1)) if len(pesos) > 1 else 0.0
        else:
            average, minimum, maximum, std_dev = None, None, None, None

        # Calcular errores T1 y T2
        rango_T1_min = net_weight - 2 * error_T1  # Límite inferior de T1
        rango_T1_max = net_weight - error_T1      # Límite superior de T1
        rango_T2 = net_weight - 2 * error_T1      # Límite inferior de T2

        # Contar unidades con error T1 y T2
        count_T1 = sum(rango_T1_min <= p < rango_T1_max for p in pesos)
        count_T2 = sum(p < rango_T2 for p in pesos)

        # Porcentaje de unidades con error T1
        percent_T1 = (count_T1 / len(pesos)) * 100 if pesos else 0

        # Determinar el resultado de aceptación del lote
        result = "Aceptado"
        if average < net_weight:
            result = "Rechazado - Promedio insuficiente"
        elif percent_T1 > 2.5:
            result = "Rechazado - Demasiados errores T1"
        elif count_T2 > 0:
            result = "Rechazado - Error T2 presente"

        # Conectar a la base de datos
        conn = get_db_connection()
        cur = conn.cursor()

        # Insertar los datos en la base de datos
        query = """
            INSERT INTO weight_control (
                date, baler, net_weight, format, brand, lot, 
                manufacture_date, expiry_date, average, minimum, maximum, standard_deviation,
                p1, p2, p3, p4, p5, p6, p7, p8, p9, p10, p11, p12, p13, p14, p15, 
                p16, p17, p18, p19, p20, p21, p22, p23, p24, p25, p26, p27, p28, p29, p30,
                count_T1, count_T2, percent_T1, result, limite_maximo_operativo, limite_minimo_operativo, ean13
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                      %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                      %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                      %s, %s, %s, %s, %s, %s, %s, %s)
        """
        values = [
            date, baler, net_weight, format_, brand, lot, manufacture_date, expiry_date,
            average, minimum, maximum, std_dev
        ] + pesos + [None] * (30 - len(pesos)) + [count_T1, count_T2, percent_T1, result, limite_maximo_operativo, limite_minimo_operativo, ean13]

        cur.execute(query, values)
        conn.commit()

        # Cerrar la conexión
        cur.close()
        conn.close()

        return jsonify({
            "message": "Registro guardado exitosamente",
            "average": average,
            "min": minimum,
            "max": maximum,
            "std_dev": std_dev,
            "count_T1": count_T1,
            "count_T2": count_T2,
            "percent_T1": percent_T1,
            "result": result,
            "limite_maximo_operativo": limite_maximo_operativo,
            "limite_minimo_operativo": limite_minimo_operativo,
            "ean13": ean13
        }), 201

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get-last-weight-summary', methods=['GET']) #Obtener la información de la ultimo ingreso de pesos
def get_last_weight_summary():

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("""
            SELECT date, baler, net_weight, format, brand, lot, 
                   manufacture_date, expiry_date, average, minimum, maximum, standard_deviation,  
                   count_t1, count_t2, 
                   limite_maximo_operativo, limite_minimo_operativo, result, ean13
            FROM weight_control 
            ORDER BY id DESC 
            LIMIT 1
        """)
        last_record = cur.fetchone()

        cur.close()
        conn.close()

        if last_record:
            return jsonify(last_record), 200
        else:
            return jsonify({"message": "No hay registros"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get-weight-history', methods=['GET']) # Obtener los datos para el grafico de dispersion | Control de pesos
def get_weight_history():

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Obtener el último registro
        cur.execute("""
            SELECT net_weight, p1, p2, p3, p4, p5, p6, p7, p8, p9, p10,
                   p11, p12, p13, p14, p15, p16, p17, p18, p19, p20,
                   p21, p22, p23, p24, p25, p26, p27, p28, p29, p30,
                   limite_maximo_operativo, limite_minimo_operativo, average
            FROM weight_control
            ORDER BY id DESC
            LIMIT 1
        """)
        record = cur.fetchone()

        cur.close()
        conn.close()

        if record:
            # Extraer los pesos en una lista
            weights = [{"x": i+1, "y": record[f"p{i+1}"]} for i in range(30)]

            return jsonify({
                "net_weight": record["net_weight"],
                "weights": weights,
                "limite_maximo_operativo": record["limite_maximo_operativo"],
                "limite_minimo_operativo": record["limite_minimo_operativo"],
                "average": record["average"]
            }), 200
        else:
            return jsonify({"message": "No hay registros"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/get-last-weights', methods=['GET']) # Obtener los ultimos pesos ingresados
def get_last_weights():
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Obtener el último registro de pesos
        cur.execute("""
            SELECT p1, p2, p3, p4, p5, p6, p7, p8, p9, p10,
                   p11, p12, p13, p14, p15, p16, p17, p18, p19, p20,
                   p21, p22, p23, p24, p25, p26, p27, p28, p29, p30
            FROM weight_control
            ORDER BY id DESC
            LIMIT 1
        """)
        record = cur.fetchone()

        cur.close()
        conn.close()

        if record:
            # Convertir los pesos en una lista simple
            weights = [record[f"p{i+1}"] for i in range(30)]

            return jsonify({"weights": weights}), 200
        else:
            return jsonify({"message": "No hay registros"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/weight-control', methods=['GET'])
def get_weight_control_data():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        query = """
        SELECT id, date, baler, net_weight, format, brand, lot, manufacture_date, expiry_date, 
               average, minimum, maximum, standard_deviation, 
               p1, p2, p3, p4, p5, p6, p7, p8, p9, p10, 
               p11, p12, p13, p14, p15, p16, p17, p18, p19, p20, 
               p21, p22, p23, p24, p25, p26, p27, p28, p29, p30, 
               result, count_t1, count_t2, percent_t1, 
               limite_maximo_operativo, limite_minimo_operativo, ean13
        FROM weight_control
        ORDER BY date DESC
        """
        cursor.execute(query)
        rows = cursor.fetchall()

        # Obtener nombres de columnas
        column_names = [desc[0] for desc in cursor.description]

        # Convertir los datos a lista de diccionarios
        data = [dict(zip(column_names, row)) for row in rows]

        # Formatear las fechas a 'DD-MM-YYYY'
        for row in data:
            if row.get('date'):
                row['date'] = row['date'].strftime('%d-%m-%Y')  # Formato de la fecha
            if row.get('manufacture_date'):
                row['manufacture_date'] = row['manufacture_date'].strftime('%d-%m-%Y')
            if row.get('expiry_date'):
                row['expiry_date'] = row['expiry_date'].strftime('%d-%m-%Y')

        cursor.close()
        conn.close()

        return jsonify(data)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/weight-control/<int:id>', methods=['PUT'])  # Editar registro de control de peso
def update_weight_control(id):
    data = request.json
    connection = get_db_connection()
    cursor = connection.cursor()

    # Generar los campos dinámicamente para la actualización
    update_fields = ", ".join([f"{key} = %s" for key in data.keys()])
    values = list(data.values()) + [id]

    # Consulta de actualización
    query = f"UPDATE weight_control SET {update_fields} WHERE id = %s;"
    
    cursor.execute(query, values)
    connection.commit()
    cursor.close()
    connection.close()

    return jsonify({"message": "Registro de peso actualizado exitosamente."})

@app.route('/weight-control/<int:id>', methods=['DELETE'])  # Eliminar registro de control de peso
def delete_weight_control(id):
    connection = get_db_connection()
    cursor = connection.cursor()

    # Ejecutar eliminación
    cursor.execute("DELETE FROM weight_control WHERE id = %s;", (id,))
    connection.commit()

    cursor.close()
    connection.close()

    return jsonify({"message": "Registro de peso eliminado exitosamente."})

@app.route('/download-weight-control', methods=['GET'])  # Descargar registro | Control de peso
def download_weight_control_excel():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute('''
            SELECT id, date, baler, net_weight, format, brand, lot, manufacture_date, expiry_date, 
                   average, minimum, maximum, standard_deviation, 
                   p1, p2, p3, p4, p5, p6, p7, p8, p9, p10, 
                   p11, p12, p13, p14, p15, p16, p17, p18, p19, p20, 
                   p21, p22, p23, p24, p25, p26, p27, p28, p29, p30, 
                   result, count_t1, count_t2, percent_t1, 
                   limite_maximo_operativo, limite_minimo_operativo, ean13
            FROM weight_control
            ORDER BY date DESC
        ''')
        data = cur.fetchall()

        # Crear el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Control de Peso"

        # Encabezados
        headers = [
            "ID", "Fecha", "Empacador", "Peso Neto", "Formato", "Marca", "Lote",
            "Fecha de Fabricación", "Fecha de Vencimiento",
            "Promedio", "Mínimo", "Máximo", "Desviación Estándar",
        ]
        headers += [f"P{i}" for i in range(1, 31)]
        headers += [
            "Resultado", "Cantidad T1", "Cantidad T2", "Porcentaje T1",
            "Límite Máx. Operativo", "Límite Mín. Operativo", "EAN13"
        ]
        ws.append(headers)

        # Función para convertir fechas a string
        def safe_str(value):
            if isinstance(value, (datetime, date)):
                return value.strftime('%d-%m-%Y')
            return value

        # Agregar filas de datos
        for row in data:
            formatted_row = [safe_str(value) for value in row]
            ws.append(formatted_row)

        cur.close()
        conn.close()

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="registro_pesos.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


#PAGINA PARA LISTAR ARTICULOS

# Obtener todos los artículos
@app.route('/articles', methods=['GET'])
def get_articles():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM articles ORDER BY id;")
    rows = cur.fetchall()
    colnames = [desc[0] for desc in cur.description]
    data = [dict(zip(colnames, row)) for row in rows]
    cur.close()
    conn.close()
    return jsonify(data)

# Insertar nuevo artículo
@app.route('/articles', methods=['POST'])
def add_article():
    data = request.json
    query = """
        INSERT INTO articles (cod_article, article_name, format, brand_id, ean13, ean14, weight)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """
    values = (
        data['cod_article'], data['article_name'], data['format'],
        data['brand_id'], data['ean13'], data['ean14'], data['weight']
    )
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, values)
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'message': 'Artículo añadido correctamente'})

# Actualizar artículo
@app.route('/articles/<int:id>', methods=['PUT'])
def update_article(id):
    data = request.json
    query = """
        UPDATE articles SET cod_article=%s, article_name=%s, format=%s,
        brand_id=%s, ean13=%s, ean14=%s, weight=%s WHERE id=%s
    """
    values = (
        data['cod_article'], data['article_name'], data['format'],
        data['brand_id'], data['ean13'], data['ean14'], data['weight'], id
    )
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, values)
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'message': 'Artículo actualizado correctamente'})

# Eliminar artículo
@app.route('/articles/<int:id>', methods=['DELETE'])
def delete_article(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM articles WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'message': 'Artículo eliminado correctamente'})

@app.route('/brands', methods=['GET'])
def get_brands():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, brand_name FROM brand ORDER BY id")
    rows = cur.fetchall()
    brands = [{'id': row[0], 'name': row[1]} for row in rows]
    cur.close()
    return jsonify(brands)


#CONTROL DE HUMEDADES

@app.route('/submit-humidity-control', methods=['POST']) # Ingresar el control realizados
def add_humidity():
    data = request.get_json()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO humidity_control (date, time, line, format, zone, humidity, responsible, observations, balance)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        data['date'], data['time'], data['line'], data['format'],
        data['zone'], data['humidity'], data['responsible'], data['observations'], data['balance']
    ))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'message': 'Humedad registrada correctamente'}), 201

@app.route('/humidity-records', methods=['GET']) # Obtener el registro de control de humeades para la tabla
def get_humidity_records():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM humidity_control ORDER BY id DESC")
        rows = cur.fetchall()
        records = []

        for row in rows:
            record = {}
            for i, value in enumerate(row):
                column_name = cur.description[i][0]
                # Convert date and time to string
                if isinstance(value, (date, time)):
                    record[column_name] = value.isoformat()
                else:
                    record[column_name] = value
            records.append(record)

        return jsonify(records)
    except Exception as e:
        print(f"❌ Error en get_humidity_records: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/get-lines', methods=['GET']) # Obtener los nombres de las lineas de producción | Control de Humedades
def get_lines():
    try:
        # Conectar a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Consulta para obtener todas las empacadoras
        query = "SELECT id, line_name FROM production_line"
        cursor.execute(query)
        lines = cursor.fetchall()

        # Cerrar conexión
        cursor.close()
        conn.close()

        return jsonify(lines), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/get-formats/<int:line_id>', methods=['GET'])# Obtener los formatos del producto por cada linea de produccións
def get_formats_by_line(line_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, format_name FROM pasta_format WHERE line_id = %s", (line_id,))
    formats = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([{"id": f[0], "format_name": f[1]} for f in formats])

@app.route('/download-humidity', methods=['GET'])  # Endpoint para descarga de registro de humedades
def download_humidity():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Consulta de los datos de la tabla humidity_control
        cur.execute('''
            SELECT date, time, line, format, zone, humidity, responsible, observations
            FROM humidity_control
            ORDER BY date DESC, time DESC
        ''')
        data = cur.fetchall()

        # Crear el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Control de Humedades"

        # Encabezados
        headers = [
            "Fecha", "Hora", "Línea", "Formato", "Zona", "Humedad (%)", "Responsable", "Observaciones"
        ]
        ws.append(headers)

        # Insertar los datos
        for record in data:
            ws.append(list(record))

        cur.close()
        conn.close()

        # Guardar en memoria (BytesIO)
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="registro-humedades.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


#PAGINA PARA ALMACENAMIENTO DE ARTES (EN PROCESO)

@app.route('/get-packaging-files', methods=['GET'])
def get_packaging_files():
    try:
        base_path = 'D:/Projects/sucesores-app-data/Artes'  
        categories = ['Rollos', 'Fundas', 'Cajas', 'Sacos']  # Nombres exactos de las carpetas

        data = []

        for category in categories:
            category_path = os.path.join(base_path, category)
            if not os.path.exists(category_path):
                continue

            category_data = {
                'name': category.replace('-', ' ').title(),  # Ej: Planos-Mecanicos → Planos Mecanicos
                'path': category,  # Para íconos o identificadores
                'brands': []
            }

            for brand in os.listdir(category_path):
                brand_path = os.path.join(category_path, brand)
                if os.path.isdir(brand_path):
                    files = [
                        f for f in os.listdir(brand_path)
                        if os.path.isfile(os.path.join(brand_path, f))
                    ]
                    category_data['brands'].append({
                        'name': brand,
                        'files': files
                    })

            data.append(category_data)

        return jsonify(data), 200

    except Exception as e:
        print(f"Error inesperado: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/files/<category>/<brand>/<filename>')
def serve_packaging_file(category, brand, filename):
    base_path = 'D:/Projects/sucesores-app-data/Artes'
    full_path = os.path.join(base_path, category, brand)

    return send_from_directory(full_path, filename)


#PAGINA PARA CONTROL DE IMPLEMENTOS DE LIMPIEZA (EN PROCESO)

@app.route('/register-movement', methods=['POST'])
def register_movement():
    data = request.json
    product_id = data['product_id']
    movement_date = data['date']
    area = data['area']
    income = float(data.get('income', 0))  # Por defecto es 0 si no se proporciona
    outcome = float(data.get('outcome', 0))  # Por defecto es 0 si no se proporciona
    responsible = data['responsible']
    observations = data.get('observations', '')

    # Verificamos que no se pueda tener tanto ingreso como egreso al mismo tiempo
    if income > 0 and outcome > 0:
        return jsonify({'error': 'No se puede tener ingreso y egreso al mismo tiempo'}), 400

    # Obtenemos la cantidad actual del producto
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # 1. Obtener el stock actual del producto
        cur.execute("SELECT quantity FROM cleaning_products WHERE id = %s", (product_id,))
        result = cur.fetchone()

        if not result:
            return jsonify({'error': 'Producto no encontrado'}), 404

        current_quantity = float(result[0])

        # 2. Calculamos el saldo basado en el tipo de movimiento
        if income > 0:
            new_balance = current_quantity + income
        elif outcome > 0:
            if outcome > current_quantity:
                return jsonify({'error': 'No hay suficiente stock disponible para este egreso'}), 400
            new_balance = current_quantity - outcome
        else:
            return jsonify({'error': 'Debe haber un ingreso o egreso mayor a 0'}), 400

        # 3. Insertamos el movimiento con el nuevo saldo
        cur.execute("""
            INSERT INTO cleaning_movements (
                product_id, date, area, income, outcome, balance, responsible, observations
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (product_id, movement_date, area, income, outcome, new_balance, responsible, observations))

        # 4. Actualizamos el stock del producto
        cur.execute("""
            UPDATE cleaning_products
            SET quantity = %s
            WHERE id = %s
        """, (new_balance, product_id))

        conn.commit()
        return jsonify({'message': 'Movimiento registrado correctamente', 'new_balance': new_balance})

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500

    finally:
        cur.close()
        conn.close()

@app.route('/cleaning-products', methods=['GET'])
def get_cleaning_products():
    try:
        product_type = request.args.get('type')

        conn = get_db_connection()
        cur = conn.cursor()

        if product_type:
            cur.execute('SELECT id, name FROM cleaning_products WHERE type = %s ORDER BY name ASC', (product_type,))
        else:
            cur.execute('SELECT id, name FROM cleaning_products ORDER BY name ASC')

        products = cur.fetchall()
        cur.close()
        conn.close()

        product_list = [{"id": product[0], "name": product[1]} for product in products]
        return jsonify(product_list), 200
    except Exception as e:
        return jsonify({"error": "Error al obtener productos", "message": str(e)}), 500

@app.route('/product-balance/<int:product_id>', methods=['GET'])
def get_product_balance(product_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT name, quantity FROM cleaning_products WHERE id = %s', (product_id,))
        product = cur.fetchone()
        cur.close()
        conn.close()

        if product:
            return jsonify({"id": product_id, "name": product[0], "quantity": product[1]}), 200
        else:
            return jsonify({"error": "Producto no encontrado"}), 404
    except Exception as e:
        return jsonify({"error": "Error al obtener saldo", "message": str(e)}), 500

@app.route('/cleaning-products-list', methods=['GET'])
def get_cleaning_products_list():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT id, name, quantity, type FROM cleaning_products ORDER BY name ASC')
        products = cur.fetchall()
        cur.close()
        conn.close()

        product_list = [{"id": p[0], "name": p[1], "quantity": p[2], "type": p[3]} for p in products]
        return jsonify(product_list), 200
    except Exception as e:
        return jsonify({"error": "Error al obtener productos", "message": str(e)}), 500

@app.route('/add-cleaning-product', methods=['POST'])
def create_product():
    data = request.get_json()
    name = data.get('name')
    quantity = data.get('quantity', 0)
    type = data.get('type')

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO cleaning_products (name, quantity, type) VALUES (%s, %s, %s)", (name, quantity, type))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"message": "Producto creado exitosamente"}), 201
    except Exception as e:
        return jsonify({"error": "Error al crear producto", "message": str(e)}), 500

@app.route('/edit-cleaning-product/<int:product_id>', methods=['PUT'])
def update_cleaning_product(product_id):
    data = request.get_json()
    name = data.get('name')
    quantity = data.get('quantity')
    type = data.get('type')

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("UPDATE cleaning_products SET name = %s, quantity = %s, type = %s WHERE id = %s", (name, quantity, type, product_id))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"message": "Producto actualizado exitosamente"}), 200
    except Exception as e:
        return jsonify({"error": "Error al actualizar producto", "message": str(e)}), 500

@app.route('/delete-cleaning-product/<int:product_id>', methods=['DELETE'])
def delete_cleaning_product(product_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM cleaning_products WHERE id = %s', (product_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'message': 'Producto eliminado correctamente'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download-cleaning-movements', methods=['GET'])
def download_cleaning_movements_excel():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute('''
            SELECT cm.id, cp.name AS product_name, cm.date, cm.area,
                   cm.income, cm.outcome, cm.balance, cm.observations
            FROM cleaning_movements cm
            JOIN cleaning_products cp ON cm.product_id = cp.id
            ORDER BY cm.date DESC
        ''')
        data = cur.fetchall()

        # Crear libro Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos Limpieza"

        # Encabezados
        headers = [
            "ID", "Producto", "Fecha", "Área", 
            "Ingreso", "Egreso", "Saldo", "Observaciones"
        ]
        ws.append(headers)

        # Función para formatear fechas
        def safe_str(value):
            if isinstance(value, (datetime, date)):
                return value.strftime('%d-%m-%Y')
            return value

        # Agregar filas
        for row in data:
            formatted_row = [safe_str(cell) for cell in row]
            ws.append(formatted_row)

        cur.close()
        conn.close()

        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="movimientos_limpieza.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

#PAGINA PARA LA GENERACION DE MULTAS

@app.route('/faults', methods=['POST'])
def add_fault():
    data = request.get_json()
    personnel_id = data['personnel_id']
    description = data.get('description', '')

    conn = get_db_connection()
    cur = conn.cursor()

    # Insertar nueva falta
    cur.execute(
        "INSERT INTO faults (personnel_id, description) VALUES (%s, %s) RETURNING id",
        (personnel_id, description)
    )
    new_fault_id = cur.fetchone()[0]

    # Obtener nuevas faltas activas no penalizadas
    cur.execute("""
        SELECT COUNT(*) FROM faults f
        LEFT JOIN penalties_faults pf ON f.id = pf.fault_id
        WHERE f.personnel_id = %s AND pf.fault_id IS NULL
    """, (personnel_id,))
    active_faults = cur.fetchone()[0]

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({'status': 'success', 'fault_id': new_fault_id, 'active_faults': active_faults})

@app.route('/generate-penalty', methods=['POST'])
def generate_penalty():
    data = request.get_json()
    personnel_id = data['personnel_id']

    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener 5 faltas no penalizadas
    cur.execute("""
        SELECT f.id FROM faults f
        LEFT JOIN penalties_faults pf ON f.id = pf.fault_id
        WHERE f.personnel_id = %s AND pf.fault_id IS NULL
        ORDER BY f.date ASC LIMIT 5
    """, (personnel_id,))
    active_faults = cur.fetchall()

    if len(active_faults) < 5:
        cur.close()
        conn.close()
        return jsonify({'status': 'error', 'message': 'No hay 5 faltas acumuladas'})

    # Crear multa
    cur.execute(
        "INSERT INTO penalties (personnel_id, penalty_description) VALUES (%s, %s) RETURNING id",
        (personnel_id, 'Multa por acumulación de 5 faltas')
    )
    penalty_id = cur.fetchone()[0]

    # Asociar faltas a la multa
    for fault in active_faults:
        cur.execute(
            "INSERT INTO penalties_faults (penalty_id, fault_id) VALUES (%s, %s)",
            (penalty_id, fault[0])
        )

    # Volver a contar las faltas activas después de la penalización
    cur.execute("""
        SELECT COUNT(*) FROM faults f
        LEFT JOIN penalties_faults pf ON f.id = pf.fault_id
        WHERE f.personnel_id = %s AND pf.fault_id IS NULL
    """, (personnel_id,))
    updated_active_faults = cur.fetchone()[0]

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        'status': 'success',
        'penalty_id': penalty_id,
        'active_faults': updated_active_faults  # <- nuevo valor retornado
    })

@app.route('/faults/<int:personnel_id>', methods=['GET'])
def get_faults(personnel_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener todas las faltas (para mostrar en tabla)
    cur.execute("""
        SELECT f.id, f.date, f.description
        FROM faults f
        WHERE f.personnel_id = %s
        ORDER BY f.date DESC, f.id DESC
    """, (personnel_id,))


    rows = cur.fetchall()

    # Obtener número de faltas activas no penalizadas (para muñeco y condición de multa)
    cur.execute("""
        SELECT COUNT(*) FROM faults f
        LEFT JOIN penalties_faults pf ON f.id = pf.fault_id
        WHERE f.personnel_id = %s AND pf.fault_id IS NULL
    """, (personnel_id,))
    active_faults = cur.fetchone()[0]

    cur.close()
    conn.close()

    faults = [{'id': r[0], 'date': r[1], 'description': r[2]} for r in rows]
    return jsonify({'faults': faults, 'active_faults': active_faults})

@app.route('/get-personnel-list', methods=['GET'])
def get_personnel_list():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT id, name FROM personnel ORDER by name ASC')  # Solo campos relevantes

        rows = cur.fetchall()
        personnel = [{"id": row[0], "name": row[1]} for row in rows]

        cur.close()
        conn.close()

        return jsonify({"personnel": personnel}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download-faults-penalties', methods=['GET'])
def download_faults_and_penalties():
    connection = get_db_connection()
    cursor = connection.cursor()

    # Consultar las faltas con nombres
    cursor.execute("""
        SELECT p.name AS employee_name, f.date, f.description
        FROM faults f
        JOIN personnel p ON f.personnel_id = p.id
        ORDER BY f.date DESC;
    """)
    faults = cursor.fetchall()

    # Consultar las multas con nombres
    cursor.execute("""
        SELECT p.name AS employee_name, m.date, m.penalty_description
        FROM penalties m
        JOIN personnel p ON m.personnel_id = p.id
        ORDER BY m.date DESC;
    """)
    penalties = cursor.fetchall()

    cursor.close()
    connection.close()

    # Crear archivo Excel en memoria
    workbook = Workbook()
    faults_sheet = workbook.active
    faults_sheet.title = "Faltas"

    penalties_sheet = workbook.create_sheet("Multas")

    # Escribir cabeceras
    faults_sheet.append(["Empleado", "Fecha", "Descripción"])
    for row in faults:
        faults_sheet.append(row)

    penalties_sheet.append(["Empleado", "Fecha", "Descripción de la Multa"])
    for row in penalties:
        penalties_sheet.append(row)

    # Guardar en memoria
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output,
                     download_name="faltas_y_multas.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.getenv('PORT', 5000)))